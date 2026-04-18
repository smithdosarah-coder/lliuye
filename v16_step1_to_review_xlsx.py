# -*- coding: utf-8 -*-
"""V16 Step 1: labeled json → review-friendly xlsx.

列设计:
  #  label  conf  source  location  text  justification  判断(✓对/✗错/⊘SKIP)  正确标签(下拉)  备注

用户操作:
  - 默认全部"✓ 对",跳过即可
  - 看到预标错了 → 改"判断"列为"✗ 错",在"正确标签"列下拉选一个
  - 觉得这条样本坏(抽错了)→ "判断"改"⊘ SKIP"

条件格式:
  - conf < 0.7 → 行橙色(提示重点看)
  - 判断 = ✗ 错 → 行浅红
"""
import json
from pathlib import Path
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.utils import get_column_letter

SRC = Path(__file__).parent / "outputs" / "v16_labeled_elements.json"
OUT = Path(__file__).parent / "outputs" / "v16_REVIEW_ME.xlsx"

data = json.load(SRC.open(encoding="utf-8"))
elems = data["elements"]

LABEL_ORDER = ["SCAFFOLD", "FILL", "CLEAR", "REWRITE", "SLOT", "PRESERVE"]
LABEL_DESC_CH = {
    "SCAFFOLD": "骨架保留",
    "FILL":     "代码填",
    "CLEAR":    "清空",
    "REWRITE":  "LLM重写",
    "SLOT":     "占位槽",
    "PRESERVE": "指引保留",
}

# 按 label 分组,然后每组内按 confidence 升序(低置信先看)
by_label = defaultdict(list)
for e in elems:
    by_label[e["label"]].append(e)
for lab in by_label:
    by_label[lab].sort(key=lambda x: x["confidence"])

# 合并:按 LABEL_ORDER 顺序
ordered_elems = []
for lab in LABEL_ORDER:
    ordered_elems.extend(by_label.get(lab, []))

wb = Workbook()
ws = wb.active
ws.title = "V16 标注 review"

headers = ["#", "我的预标", "置信", "源文件", "位置", "文本", "规则说明",
           "你的判断", "如错,正确标签", "备注"]
ws.append(headers)

# 表头样式
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill("solid", fgColor="2E5D82")
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left = Alignment(horizontal="left", vertical="center", wrap_text=True)
thin = Side(border_style="thin", color="CCCCCC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for ci, cell in enumerate(ws[1], start=1):
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center
    cell.border = border

# 数据行
for i, e in enumerate(ordered_elems, start=1):
    row_idx = i + 1  # 行号(1 起,首行是 header)
    lab = e["label"]
    conf = float(e["confidence"])
    src_short = e["source"].replace("_对公成稿", "").replace("_骨架型", "")[:14]
    text = (e["text"] or "")[:180]
    just = (e["justification"] or "")[:100]

    ws.cell(row=row_idx, column=1, value=i)
    ws.cell(row=row_idx, column=2, value=f"{lab} · {LABEL_DESC_CH[lab]}")
    ws.cell(row=row_idx, column=3, value=round(conf, 2))
    ws.cell(row=row_idx, column=4, value=src_short)
    ws.cell(row=row_idx, column=5, value=e["location"])
    ws.cell(row=row_idx, column=6, value=text)
    ws.cell(row=row_idx, column=7, value=just)
    ws.cell(row=row_idx, column=8, value="✓ 对")  # 默认"对"
    # 列 9 (正确标签) 默认空
    # 列 10 (备注) 默认空

    for c in range(1, 11):
        cell = ws.cell(row=row_idx, column=c)
        cell.border = border
        cell.alignment = left if c in (6, 7, 10) else center

# 列宽
widths = {1: 5, 2: 16, 3: 7, 4: 14, 5: 16, 6: 70, 7: 40, 8: 12, 9: 16, 10: 20}
for idx, w in widths.items():
    ws.column_dimensions[get_column_letter(idx)].width = w

# 冻结首行
ws.freeze_panes = "A2"

# 行高 — 默认给 30 让 wrap_text 生效
for r in range(2, len(ordered_elems) + 2):
    ws.row_dimensions[r].height = 30

# 数据验证 1: "你的判断" 列 (H 列, 第 8 列)
dv_judge = DataValidation(
    type="list",
    formula1='"✓ 对,✗ 错,⊘ SKIP"',
    allow_blank=False,
    showDropDown=False,  # False 意味着显示下拉箭头(openpyxl 反义)
)
dv_judge.error = "请从下拉选择"
dv_judge.errorTitle = "只能选 ✓ 对 / ✗ 错 / ⊘ SKIP"
dv_judge.prompt = "默认 ✓ 对。预标错了改成 ✗ 错,样本本身坏选 ⊘ SKIP"
dv_judge.promptTitle = "对预标的判断"
dv_judge.add(f"H2:H{len(ordered_elems) + 1}")
ws.add_data_validation(dv_judge)

# 数据验证 2: "正确标签" 列 (I 列, 第 9 列)
correct_labels = ",".join(LABEL_ORDER)
dv_correct = DataValidation(
    type="list",
    formula1=f'"{correct_labels}"',
    allow_blank=True,
    showDropDown=False,
)
dv_correct.prompt = "仅在判断列选 ✗ 错 时需要填,从下拉选正确标签"
dv_correct.promptTitle = "正确标签"
dv_correct.add(f"I2:I{len(ordered_elems) + 1}")
ws.add_data_validation(dv_correct)

# 条件格式 1: 置信度 < 0.7 → 行橙色提醒
orange_fill = PatternFill("solid", fgColor="FFE4B5")
rule_low_conf = FormulaRule(
    formula=[f"$C2<0.7"],
    fill=orange_fill,
)
ws.conditional_formatting.add(f"A2:J{len(ordered_elems) + 1}", rule_low_conf)

# 条件格式 2: 判断 = ✗ 错 → 行浅红
light_red = PatternFill("solid", fgColor="FFDDDD")
rule_wrong = FormulaRule(
    formula=[f'$H2="✗ 错"'],
    fill=light_red,
)
ws.conditional_formatting.add(f"A2:J{len(ordered_elems) + 1}", rule_wrong)

# 条件格式 3: 判断 = ⊘ SKIP → 行灰
light_grey = PatternFill("solid", fgColor="DDDDDD")
rule_skip = FormulaRule(
    formula=[f'$H2="⊘ SKIP"'],
    fill=light_grey,
)
ws.conditional_formatting.add(f"A2:J{len(ordered_elems) + 1}", rule_skip)

# 额外 sheet: 使用说明
ws2 = wb.create_sheet("使用说明")
help_lines = [
    "V16 Step 1 标注 Review — 使用说明",
    "",
    "当前假想客户:福建中锐网络股份有限公司(黄祖海实控,智慧水利+智慧教育,4100 万)",
    "",
    "【操作流程】",
    "1. 默认 '你的判断' 列已填 '✓ 对' — 意思是我的预标没问题,跳过",
    "2. 看到预标错了 → 把 '你的判断' 改成 '✗ 错',然后在 '如错,正确标签' 列下拉选一个",
    "3. 如果这条样本根本不该抽出来(坏样本)→ '你的判断' 选 '⊘ SKIP'",
    "4. 关键 review 点:置信度 <0.7 的行(橙色)- 这些是规则不确定的,重点看",
    "5. 全部过完保存关闭,把文件发我(或告诉我路径)",
    "",
    "【6 类标签含义】",
    "SCAFFOLD  骨架保留原文(章节标题 / 字段标签 / 表头 / 单位:万元 / 指引)",
    "FILL      有当前客户数据,代码直接填(注册资本 / 法人 / 公司名)",
    "CLEAR     是示例但当前客户无对应数据,清空+pending(PD评级 / 申报单位)",
    "REWRITE   正文段落,LLM 基于当前客户材料整段重写(经营分析 / 财务描述)",
    "SLOT      占位槽位(下划线 / 空白 / '(面积等)')",
    "PRESERVE  说明性指引保留(仅兜底,实际用不多)",
    "",
    "【预估时间】10-15 分钟,192 条样本平均 5 秒/条",
]
for i, ln in enumerate(help_lines, 1):
    ws2.cell(row=i, column=1, value=ln)
    if i == 1:
        ws2.cell(row=i, column=1).font = Font(bold=True, size=14)
ws2.column_dimensions["A"].width = 80

wb.save(OUT)
print(f"[done] {OUT}")
print(f"  总样本: {len(ordered_elems)}")
print(f"  低置信(橙色)需重点 review: {sum(1 for e in ordered_elems if e['confidence'] < 0.7)}")
