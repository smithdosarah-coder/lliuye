# -*- coding: utf-8 -*-
"""Exporter: HitList → Excel/Word/Markdown 导出。"""

from __future__ import annotations
from pathlib import Path
from datetime import datetime

from .models import HitList, RiskLevel


def export_hitlist_excel(hl: HitList, output_dir: str | Path = "outputs") -> str:
    """导出分级命中榜为 Excel。"""
    out_dir = Path(output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = out_dir / f"{hl.agent_name}_hitlist_{ts}.xlsx"

    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment
    except Exception:
        # 降级为 CSV
        return export_hitlist_csv(hl, output_dir)

    wb = Workbook()
    ws = wb.active
    ws.title = "命中榜"

    headers = ["排名", "分级", "分数", "目标ID", "目标名称", "命中规则", "理由", "证据摘要"]
    ws.append(headers)

    level_fill = {
        RiskLevel.RED: PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid"),
        RiskLevel.YELLOW: PatternFill(start_color="FFF9E6", end_color="FFF9E6", fill_type="solid"),
        RiskLevel.GREEN: PatternFill(start_color="E6F9E6", end_color="E6F9E6", fill_type="solid"),
    }
    level_label = {
        RiskLevel.RED: "🔴 红",
        RiskLevel.YELLOW: "🟡 黄",
        RiskLevel.GREEN: "🟢 绿",
        RiskLevel.INFO: "ℹ️ 信息",
    }

    for h in hl.hits:
        target_name = h.target.payload.get("company_name") or h.target.payload.get("title") or h.target.target_id
        evidence_summary = " | ".join(e.snippet[:60] for e in h.evidences[:3])
        ws.append([
            h.rank,
            level_label.get(h.level, str(h.level)),
            round(h.score, 2),
            h.target.target_id,
            target_name,
            ", ".join(h.matched_rules),
            " | ".join(h.reasons[:3]),
            evidence_summary,
        ])
        # 行染色
        fill = level_fill.get(h.level)
        if fill:
            for c in ws[ws.max_row]:
                c.fill = fill

    # 列宽
    widths = [6, 10, 8, 18, 30, 25, 50, 60]
    for i, w in enumerate(widths):
        ws.column_dimensions[chr(ord("A") + i)].width = w

    # 顶部汇总
    ws.insert_rows(1, amount=3)
    ws["A1"] = f"{hl.agent_name} 分级命中榜"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"生成时间：{hl.generated_at} | KB：{hl.kb_summary} | 扫描：{hl.scan_summary}"
    ws["A3"] = f"共扫描 {hl.total_scanned} 目标 · 🔴 {hl.red_count}  🟡 {hl.yellow_count}  🟢 {hl.green_count}"

    wb.save(str(path))
    return str(path)


def export_hitlist_csv(hl: HitList, output_dir: str | Path = "outputs") -> str:
    import csv
    out_dir = Path(output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = out_dir / f"{hl.agent_name}_hitlist_{ts}.csv"
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        w.writerow(["排名", "分级", "分数", "目标ID", "目标名称", "命中规则", "理由"])
        for h in hl.hits:
            target_name = h.target.payload.get("company_name") or h.target.payload.get("title") or h.target.target_id
            w.writerow([
                h.rank, h.level.value, round(h.score, 2), h.target.target_id,
                target_name, ", ".join(h.matched_rules), " | ".join(h.reasons[:3]),
            ])
    return str(path)


def export_hitlist_markdown(hl: HitList, output_dir: str | Path = "outputs") -> str:
    out_dir = Path(output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = out_dir / f"{hl.agent_name}_hitlist_{ts}.md"

    lines = [
        f"# {hl.agent_name} 分级命中榜",
        f"",
        f"- 生成时间：{hl.generated_at}",
        f"- 知识库：{hl.kb_summary}",
        f"- 扫描范围：{hl.scan_summary}",
        f"- 共扫描 **{hl.total_scanned}** 个目标",
        f"- 🔴 严重 **{hl.red_count}**  ·  🟡 一般 **{hl.yellow_count}**  ·  🟢 观察 **{hl.green_count}**",
        "",
        "---",
        "",
    ]
    label = {
        RiskLevel.RED: "🔴 红",
        RiskLevel.YELLOW: "🟡 黄",
        RiskLevel.GREEN: "🟢 绿",
        RiskLevel.INFO: "ℹ️ 信息",
    }
    for h in hl.hits:
        target_name = h.target.payload.get("company_name") or h.target.payload.get("title") or h.target.target_id
        lines.append(f"### #{h.rank} [{label.get(h.level)}] {target_name} ({round(h.score,1)}分)")
        if h.matched_rules:
            lines.append(f"- 命中规则：`{', '.join(h.matched_rules)}`")
        for r in h.reasons[:5]:
            lines.append(f"- 理由：{r}")
        for e in h.evidences[:3]:
            lines.append(f"  - 证据 @ {e.source}: {e.snippet[:120]}")
        lines.append("")

    path.write_text("\n".join(lines), encoding="utf-8")
    return str(path)
