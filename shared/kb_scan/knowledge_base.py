# -*- coding: utf-8 -*-
"""KnowledgeBase: 用户上传文件 → 结构化规则/客户/条款。

三个 Agent 的 KB 差异在于"装什么"：
  Agent1: 已有客户名录 + 政策 + 行业指引
  Agent4: 预警规则 + 在贷客户名录 + 管理制度
  Agent5: 政策库 + 内部制度 + 业务数据
"""

from __future__ import annotations
import json
import csv
from pathlib import Path

from .models import RuleItem, CompanyProfile, RuleType


SUPPORTED_EXT = {".xlsx", ".xls", ".csv", ".json", ".jsonl", ".docx", ".doc", ".pdf", ".md", ".txt"}


class KnowledgeBase:
    """通用知识库容器。"""

    def __init__(self, name: str = ""):
        self.name = name
        self.rules: list[RuleItem] = []
        self.companies: list[CompanyProfile] = []
        self.clauses: list[dict] = []          # 政策条款（Agent5 用）
        self.business_events: list[dict] = []  # 业务数据（Agent5 用）
        self.raw_files: list[dict] = []

    # ---- 装载入口 ----

    def load_file(self, file_path: str, doc_type: str = "auto") -> dict:
        path = Path(file_path)
        if not path.exists():
            return {"file": str(path), "status": "not_found"}
        if doc_type == "auto":
            doc_type = self._infer_type(path)
        meta = {"path": str(path), "type": doc_type, "status": "ok"}
        try:
            if doc_type == "customer_list":
                added = self._parse_customer_list(path)
                self.companies.extend(added)
                meta["added_companies"] = len(added)
            elif doc_type == "rule":
                added = self._parse_rule_doc(path)
                self.rules.extend(added)
                meta["added_rules"] = len(added)
            elif doc_type == "policy":
                clauses = self._parse_policy_doc(path)
                self.clauses.extend(clauses)
                meta["added_clauses"] = len(clauses)
            elif doc_type == "business_data":
                events = self._parse_business_data(path)
                self.business_events.extend(events)
                meta["added_events"] = len(events)
            elif doc_type == "industry_guide":
                meta["note"] = "industry_guide 存原文供 LLM 参考"
                meta["raw_text"] = self._read_text(path)[:5000]
            else:
                meta["status"] = "skipped"
                meta["reason"] = f"unknown doc_type={doc_type}"
        except (OSError, RuntimeError, ValueError, TypeError, AttributeError, KeyError, ImportError) as e:
            meta["status"] = "error"
            meta["error"] = str(e)[:200]
        self.raw_files.append(meta)
        return meta

    def load_files(self, file_paths: list[str]) -> list[dict]:
        return [self.load_file(fp) for fp in file_paths]

    # ---- 查询 ----

    def get_rules(self, category: str | None = None) -> list[RuleItem]:
        if not category:
            return self.rules
        return [r for r in self.rules if r.category == category]

    def get_companies(self) -> list[CompanyProfile]:
        return self.companies

    def get_clauses(self) -> list[dict]:
        return self.clauses

    def get_business_events(self) -> list[dict]:
        return self.business_events

    def summary(self) -> str:
        parts = []
        if self.companies:
            parts.append(f"{len(self.companies)} 家客户名录")
        if self.rules:
            parts.append(f"{len(self.rules)} 条规则")
        if self.clauses:
            parts.append(f"{len(self.clauses)} 条政策条款")
        if self.business_events:
            parts.append(f"{len(self.business_events)} 条业务事件")
        return " + ".join(parts) or "空知识库"

    # ---- 类型推断 ----

    def _infer_type(self, path: Path) -> str:
        name = path.name.lower()
        if any(k in name for k in ["客户", "customer", "名录", "list"]):
            return "customer_list"
        if any(k in name for k in ["规则", "rule", "预警"]):
            return "rule"
        if any(k in name for k in ["政策", "办法", "通知", "policy", "regulation"]):
            return "policy"
        if any(k in name for k in ["业务", "放款", "台账", "流水", "ledger", "transaction"]):
            return "business_data"
        if any(k in name for k in ["行业", "industry", "指引", "guide"]):
            return "industry_guide"
        # 按扩展名兜底
        if path.suffix.lower() in {".xlsx", ".xls", ".csv"}:
            return "customer_list"
        if path.suffix.lower() in {".json", ".jsonl"}:
            return "rule"
        return "policy"

    # ---- 解析器 ----

    def _read_text(self, path: Path) -> str:
        ext = path.suffix.lower()
        if ext in {".md", ".txt"}:
            return path.read_text(encoding="utf-8", errors="ignore")
        if ext == ".docx":
            try:
                from docx import Document
                doc = Document(str(path))
                return "\n".join(p.text for p in doc.paragraphs if p.text.strip())
            except (OSError, ImportError, ValueError, TypeError, AttributeError, KeyError):
                return ""
        if ext == ".pdf":
            try:
                import pdfplumber
                texts = []
                with pdfplumber.open(str(path)) as pdf:
                    for page in pdf.pages:
                        t = page.extract_text() or ""
                        texts.append(t)
                return "\n".join(texts)
            except (OSError, ImportError, ValueError, TypeError, AttributeError, KeyError):
                return ""
        return ""

    def _parse_customer_list(self, path: Path) -> list[CompanyProfile]:
        """Excel/CSV/JSON/JSONL → CompanyProfile 列表。"""
        ext = path.suffix.lower()
        rows: list[dict] = []
        if ext in {".xlsx", ".xls"}:
            try:
                from openpyxl import load_workbook
                wb = load_workbook(str(path), read_only=True, data_only=True)
                ws = wb.active
                header = None
                for row in ws.iter_rows(values_only=True):
                    if header is None:
                        header = [str(c).strip() if c is not None else f"col{i}"
                                  for i, c in enumerate(row)]
                        continue
                    rows.append({header[i]: row[i] for i in range(min(len(header), len(row)))})
            except (OSError, ImportError, ValueError, KeyError, TypeError, AttributeError):
                pass
        elif ext == ".csv":
            with path.open("r", encoding="utf-8", errors="ignore") as f:
                for r in csv.DictReader(f):
                    rows.append(r)
        elif ext == ".json":
            data = json.loads(path.read_text(encoding="utf-8", errors="ignore"))
            if isinstance(data, list):
                rows = data
        elif ext == ".jsonl":
            for line in path.read_text(encoding="utf-8", errors="ignore").splitlines():
                line = line.strip()
                if line:
                    try:
                        rows.append(json.loads(line))
                    except (json.JSONDecodeError, ValueError, TypeError):
                        pass

        out: list[CompanyProfile] = []
        for r in rows:
            if not isinstance(r, dict):
                continue
            name = r.get("company_name") or r.get("企业名称") or r.get("客户名称") or r.get("name")
            if not name:
                continue
            try:
                out.append(CompanyProfile(
                    company_id=str(r.get("company_id") or r.get("客户号") or ""),
                    company_name=str(name),
                    unified_credit_code=str(r.get("unified_credit_code") or r.get("统一社会信用代码") or ""),
                    industry=str(r.get("industry") or r.get("行业") or ""),
                    sub_industry=str(r.get("sub_industry") or r.get("细分行业") or ""),
                    region=str(r.get("region") or r.get("地区") or r.get("所在地") or ""),
                    scale=str(r.get("scale") or r.get("规模") or ""),
                    revenue_latest=str(r.get("revenue_latest") or r.get("营收") or ""),
                    employee_count=int(r.get("employee_count") or r.get("员工数") or 0) if str(
                        r.get("employee_count") or r.get("员工数") or "0").replace(".", "").isdigit() else 0,
                    keywords=[k for k in (r.get("keywords") or "").split(",") if k] if isinstance(r.get("keywords"), str) else (r.get("keywords") or []),
                    qualifications=(r.get("qualifications") or []) if isinstance(r.get("qualifications"), list) else [k for k in str(r.get("qualifications") or "").split(",") if k],
                    tags=(r.get("tags") or []) if isinstance(r.get("tags"), list) else [k for k in str(r.get("tags") or "").split(",") if k],
                    credit_balance=str(r.get("credit_balance") or r.get("在贷余额") or ""),
                    overdue_days=int(r.get("overdue_days") or r.get("逾期天数") or 0) if str(
                        r.get("overdue_days") or r.get("逾期天数") or "0").replace(".", "").isdigit() else 0,
                    extras={k: v for k, v in r.items() if k not in {"company_name", "企业名称"}},
                ))
            except (ValueError, TypeError, KeyError, AttributeError):
                continue
        return out

    def _parse_rule_doc(self, path: Path) -> list[RuleItem]:
        """规则文档 → RuleItem。结构化的 JSON 直接解析；非结构化留给 RuleExtractor。"""
        ext = path.suffix.lower()
        if ext in {".json", ".jsonl"}:
            return self._parse_rule_json(path)
        # 非结构化规则（Word/PDF）— 交给 RuleExtractor 按需处理
        # 此处只占位：保留原文，调用方自行触发 extraction
        text = self._read_text(path)
        if not text:
            return []
        # 最简单的"一段一规则"兜底分段（避免空手回）
        rules = []
        for idx, para in enumerate(text.split("\n")):
            para = para.strip()
            if len(para) < 15:
                continue
            rules.append(RuleItem(
                rule_id=f"{path.stem}_P{idx:03d}",
                source_doc=path.name,
                category="未分类",
                title=para[:40],
                content=para[:500],
                rule_type=RuleType.EXTRACTED,
            ))
        return rules

    def _parse_rule_json(self, path: Path) -> list[RuleItem]:
        ext = path.suffix.lower()
        data: list = []
        if ext == ".json":
            raw = json.loads(path.read_text(encoding="utf-8", errors="ignore"))
            if isinstance(raw, list):
                data = raw
            elif isinstance(raw, dict) and "rules" in raw:
                data = raw["rules"]
        elif ext == ".jsonl":
            for line in path.read_text(encoding="utf-8", errors="ignore").splitlines():
                line = line.strip()
                if line:
                    try:
                        data.append(json.loads(line))
                    except (json.JSONDecodeError, ValueError, TypeError):
                        pass
        out = []
        for i, d in enumerate(data):
            if not isinstance(d, dict):
                continue
            try:
                d.setdefault("rule_id", d.get("id") or f"{path.stem}_R{i:03d}")
                d.setdefault("source_doc", path.name)
                d.setdefault("rule_type", RuleType.STRUCTURED)
                out.append(RuleItem(**d))
            except (ValueError, TypeError, KeyError, AttributeError):
                continue
        return out

    def _parse_policy_doc(self, path: Path) -> list[dict]:
        """政策文档 → 按章条切分的 clauses 列表。"""
        text = self._read_text(path)
        if not text:
            return []
        import re
        # 按"第X条 / 第X章 / 一、二、"切分
        pattern = r"(第[一二三四五六七八九十百千0-9]+条[^。\n]{0,60})"
        parts = re.split(pattern, text)
        clauses = []
        for i in range(1, len(parts), 2):
            title = parts[i].strip()
            body = parts[i + 1] if i + 1 < len(parts) else ""
            body = body.strip()
            if not body:
                continue
            clauses.append({
                "clause_id": f"{path.stem}_{len(clauses)+1:03d}",
                "title": title,
                "content": body[:1500],
                "source_doc": path.name,
            })
        if not clauses:
            # 兜底：整体作为一条
            clauses.append({
                "clause_id": f"{path.stem}_001",
                "title": path.stem,
                "content": text[:5000],
                "source_doc": path.name,
            })
        return clauses

    def _parse_business_data(self, path: Path) -> list[dict]:
        """业务数据 Excel/CSV → dict 列表。"""
        ext = path.suffix.lower()
        rows: list[dict] = []
        if ext in {".xlsx", ".xls"}:
            try:
                from openpyxl import load_workbook
                wb = load_workbook(str(path), read_only=True, data_only=True)
                for ws in wb.worksheets:
                    header = None
                    for row in ws.iter_rows(values_only=True):
                        if header is None:
                            header = [str(c).strip() if c is not None else f"col{i}"
                                      for i, c in enumerate(row)]
                            continue
                        row_dict = {header[i]: row[i] for i in range(min(len(header), len(row)))}
                        row_dict["_sheet"] = ws.title
                        rows.append(row_dict)
            except (OSError, ImportError, ValueError, KeyError, TypeError, AttributeError):
                pass
        elif ext == ".csv":
            with path.open("r", encoding="utf-8", errors="ignore") as f:
                rows = list(csv.DictReader(f))
        elif ext == ".json":
            data = json.loads(path.read_text(encoding="utf-8", errors="ignore"))
            if isinstance(data, list):
                rows = data
        return rows
