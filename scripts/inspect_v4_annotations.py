"""读出 v4 xlsx 的所有备注和用户标注（K/L/M 列）."""
import io
import sys
from pathlib import Path

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
from openpyxl import load_workbook

xlsx = Path(r"D:\claude code\credit_report_agent_work\outputs\v16_REVIEW_ME_v4.xlsx")

wb = load_workbook(xlsx, read_only=True, data_only=True)
ws = wb.active

print(f"sheet: {ws.title}, max_row={ws.max_row}, max_col={ws.max_column}")
print("-" * 120)
# 列: # / LLM分类 / 规则预标 / 一致? / LLM置信 / 源 / 位置 / 文本 / LLM理由 / 规则理由 / 你的判断 / 正确label / 备注
headers = []
for cell in ws[1]:
    headers.append(cell.value)
print("headers:", headers)

flagged = []
for row in ws.iter_rows(min_row=2, values_only=True):
    if len(row) < 13:
        continue
    num, llm_label, rule_label, agree, conf, src, loc, text, llm_reason, rule_reason, user_judge, correct_label, remark = row[:13]
    # 只看用户留了备注/标记/正确label 的行
    if remark or user_judge or correct_label:
        flagged.append({
            "row": num,
            "llm": llm_label,
            "rule": rule_label,
            "agree": agree,
            "conf": conf,
            "loc": loc,
            "text": text,
            "judge": user_judge,
            "correct": correct_label,
            "remark": remark,
        })

print(f"\n== flagged rows: {len(flagged)} ==")
for f in flagged:
    print(f"\n[#{f['row']}] loc={f['loc']}")
    print(f"  llm={f['llm']}, rule={f['rule']}, agree={f['agree']}, conf={f['conf']}")
    print(f"  text: {f['text']!r}")
    print(f"  judge={f['judge']!r}, correct={f['correct']!r}")
    print(f"  remark: {f['remark']!r}")
