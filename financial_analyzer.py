# -*- coding: utf-8 -*-
"""
FinancialAnalyzer — 从标准会企01/02表xlsx财务报表中直接抽取结构化指标。

设计目标:代码算所有比率/同比/异常,LLM 只做语言加工,彻底消除算术幻觉。

三张标准表的解析策略:
  资产负债表:col A 资产科目 / col C 期末 / col D 年初;col E 负债科目 / col G 期末 / col H 年初
  利润表:col A 科目 / col C 当期 / col D 上期
  现金流量表:col A 科目 / col B 当期累计(仅 1 期)

科目匹配方式:strip() 后精确匹配标准科目名(容忍全角空格/前后空白)。
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from typing import Optional

try:
    from openpyxl import load_workbook
except ImportError:
    raise ImportError("需要 openpyxl: pip install openpyxl")


# ============================================================
# 数据契约
# ============================================================

@dataclass
class PeriodValue:
    """某指标的多期值 + 对比。
    compare_caption: BS 科目用"较年初"(期末vs期初同一主体不同时点);
                     IS/CF 用"同比"(当期vs上期流量对比)。
    """
    current: Optional[float] = None
    previous: Optional[float] = None
    yoy_abs: Optional[float] = None
    yoy_pct: Optional[float] = None
    current_label: str = "2025末"
    previous_label: str = "2024末"
    unit: str = "万元"
    compare_caption: str = "同比"

    def __post_init__(self):
        if self.current is not None and self.previous is not None:
            self.yoy_abs = self.current - self.previous
            if self.previous != 0:
                self.yoy_pct = (self.current - self.previous) / abs(self.previous) * 100

    def format(self) -> str:
        if self.current is None:
            return "无数据"
        if self.unit == "%":
            cur_s = f"{self.current:.1f}%"
        elif self.unit == "倍":
            cur_s = f"{self.current:.2f}倍"
        elif self.unit == "天":
            cur_s = f"{self.current:.0f}天"
        elif self.unit == "万元":
            cur_s = f"{self.current/10000:.1f}万元"
        else:
            cur_s = f"{self.current:.2f}{self.unit}"
        if self.previous is None:
            return f"{cur_s}"
        cap = self.compare_caption
        # cap 为"较年初"时,用"较年初"前缀;"同比"则用"同比"前缀
        # 生成LLM可直接原文复用的中文变动短语,杜绝自行计算
        if self.unit == "%":
            prev_s = f"{self.previous:.1f}%"
            if self.yoy_abs is not None:
                verb = "下降" if self.yoy_abs < 0 else ("上升" if self.yoy_abs > 0 else "持平")
                if verb == "持平":
                    yoy_s = f"{cap}持平"
                else:
                    yoy_s = f"{cap}{verb}{abs(self.yoy_abs):.1f}个百分点"
            else:
                yoy_s = ""
        elif self.unit == "倍":
            prev_s = f"{self.previous:.2f}倍"
            if self.yoy_abs is not None:
                verb = "下降" if self.yoy_abs < 0 else ("提升" if self.yoy_abs > 0 else "持平")
                if verb == "持平":
                    yoy_s = f"{cap}持平"
                else:
                    yoy_s = f"{cap}{verb}{abs(self.yoy_abs):.2f}倍"
            else:
                yoy_s = ""
        elif self.unit == "天":
            prev_s = f"{self.previous:.0f}天"
            if self.yoy_abs is not None:
                verb = "缩短" if self.yoy_abs < 0 else ("延长" if self.yoy_abs > 0 else "持平")
                if verb == "持平":
                    yoy_s = f"{cap}持平"
                else:
                    yoy_s = f"{cap}{verb}{abs(self.yoy_abs):.0f}天"
            else:
                yoy_s = ""
        elif self.unit == "万元":
            prev_s = f"{self.previous/10000:.1f}万元"
            if self.yoy_pct is not None:
                verb = "下降" if self.yoy_pct < 0 else ("增长" if self.yoy_pct > 0 else "持平")
                if verb == "持平":
                    yoy_s = f"{cap}持平"
                else:
                    yoy_s = f"{cap}{verb}{abs(self.yoy_pct):.1f}%"
            else:
                yoy_s = ""
        else:
            prev_s = f"{self.previous:.2f}{self.unit}"
            if self.yoy_pct is not None:
                verb = "下降" if self.yoy_pct < 0 else ("增长" if self.yoy_pct > 0 else "持平")
                yoy_s = f"{cap}{verb}{abs(self.yoy_pct):.1f}%" if verb != "持平" else f"{cap}持平"
            else:
                yoy_s = ""
        return f"{cur_s} ({yoy_s} | {self.previous_label}:{prev_s} → {self.current_label}:{cur_s})"


@dataclass
class FinancialStatements:
    balance_sheet: dict = field(default_factory=dict)
    income_statement: dict = field(default_factory=dict)
    cash_flow: dict = field(default_factory=dict)
    source_file: str = ""
    statement_date: str = ""
    reporting_entity: str = ""


@dataclass
class FinancialIndicators:
    # 规模
    total_assets: PeriodValue = field(default_factory=PeriodValue)
    total_liabilities: PeriodValue = field(default_factory=PeriodValue)
    total_equity: PeriodValue = field(default_factory=PeriodValue)
    revenue: PeriodValue = field(default_factory=PeriodValue)
    net_profit: PeriodValue = field(default_factory=PeriodValue)
    # 盈利
    gross_margin: PeriodValue = field(default_factory=lambda: PeriodValue(unit="%"))
    net_margin: PeriodValue = field(default_factory=lambda: PeriodValue(unit="%"))
    operating_margin: PeriodValue = field(default_factory=lambda: PeriodValue(unit="%"))
    # 偿债
    debt_to_asset_ratio: PeriodValue = field(default_factory=lambda: PeriodValue(unit="%"))
    current_ratio: PeriodValue = field(default_factory=lambda: PeriodValue(unit="倍"))
    quick_ratio: PeriodValue = field(default_factory=lambda: PeriodValue(unit="倍"))
    # 营运
    ar_turnover_days: PeriodValue = field(default_factory=lambda: PeriodValue(unit="天"))
    inventory_turnover_days: PeriodValue = field(default_factory=lambda: PeriodValue(unit="天"))
    # 现金流(当期)
    operating_cf_net: Optional[float] = None
    investing_cf_net: Optional[float] = None
    financing_cf_net: Optional[float] = None
    ocf_to_net_income: Optional[float] = None
    cash_received_to_revenue: Optional[float] = None
    # 定性
    anomalies: list = field(default_factory=list)
    trend_labels: dict = field(default_factory=dict)
    raw_statements: Optional[FinancialStatements] = None


# ============================================================
# 标准科目清单(会企01/02)
# ============================================================

BS_LEFT_SUBJECTS = {
    "货币资金", "交易性金融资产", "应收票据", "应收账款",
    "预付款项", "应收利息", "应收股利", "其他应收款", "存货",
    "一年内到期的非流动资产", "其他流动资产", "流动资产合计",
    "可供出售金融资产", "持有至到期投资", "长期应收款",
    "长期股权投资", "其他权益工具投资", "固定资产", "在建工程",
    "工程物资", "固定资产清理", "生产性生物资产", "油气资产",
    "无形资产", "开发支出", "商誉", "长期待摊费用",
    "递延所得税资产", "其他非流动资产", "非流动资产合计", "资产总计",
}

BS_RIGHT_SUBJECTS = {
    "短期借款", "交易性金融负债", "应付票据", "应付账款",
    "预收款项", "应付职工薪酬", "应交税费", "应付利息",
    "应付股利", "其他应付款", "一年内到期的非流动负债",
    "其他流动负债", "流动负债合计",
    "长期借款", "应付债券", "长期应付款", "专项应付款",
    "预计负债", "递延所得税负债", "其他非流动负债",
    "非流动负债合计", "负债合计",
    "实收资本", "实收资本(或股本)", "资本公积", "减:库存股", "库存股",
    "盈余公积", "未分配利润",
    "所有者权益合计", "所有者权益(或股东权益)合计",
    "负债和所有者权益总计", "负债和所有者权益(或股东权益)总计",
}

IS_SUBJECTS = {
    "营业收入", "营业成本", "营业税金及附加", "销售费用",
    "管理费用", "研发费用", "财务费用", "利息费用", "利息收入",
    "其他收益", "投资收益", "公允价值变动收益", "信用减值损失",
    "资产减值损失", "资产处置收益",
    "营业利润", "营业外收入", "营业外支出",
    "利润总额", "所得税费用", "净利润",
}

CF_SUBJECTS = {
    "销售商品、提供劳务收到的现金", "收到的税费返还",
    "收到其他与经营活动有关的现金", "经营活动现金流入小计",
    "购买商品、接受劳务支付的现金", "支付给职工以及为职工支付的现金",
    "支付的各项税费", "支付其他与经营活动有关的现金",
    "经营活动现金流出小计", "经营活动产生的现金流量净额",
    "投资活动现金流入小计", "投资活动现金流出小计",
    "投资活动产生的现金流量净额",
    "筹资活动现金流入小计", "筹资活动现金流出小计",
    "筹资活动产生的现金流量净额",
    "汇率变动对现金及现金等价物的影响",
    "现金及现金等价物净增加额", "期末现金及现金等价物余额",
}


_CN_NUM_PREFIX = re.compile(r"^[一二三四五六七八九十]+[、\.]")
_LIST_PREFIX = re.compile(r"^(?:减|加|其中)[:：]")
_PAREN_NOTE = re.compile(r"[\(（][^)）]*[\)）]")


def _norm(s):
    """规范化科目名:去全角空白/冒号/前缀/括号注释。"""
    if s is None:
        return ""
    s = str(s).strip()
    # 全角空格/制表
    s = s.replace("\u3000", "").replace("\t", "").replace(" ", "")
    # 中英文标点标准化
    s = s.replace("：", ":").replace("（", "(").replace("）", ")")
    # 去"一、/二、..."中文数字前缀
    s = _CN_NUM_PREFIX.sub("", s)
    # 去"减:/加:/其中:"列表前缀
    s = _LIST_PREFIX.sub("", s)
    # 去尾部括号注释,如"(亏损以"- "号填列)"
    s = _PAREN_NOTE.sub("", s)
    return s.strip()


def _parse_num(v):
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    try:
        s = str(v).replace(",", "").strip()
        if not s:
            return None
        return float(s)
    except (ValueError, TypeError):
        return None


# ============================================================
# FinancialAnalyzer
# ============================================================

class FinancialAnalyzer:

    # ------- xlsx 解析 -------

    def parse_xlsx(self, file_path: str) -> FinancialStatements:
        wb = load_workbook(file_path, data_only=True)
        stmts = FinancialStatements(source_file=file_path)

        for sname in wb.sheetnames:
            ws = wb[sname]
            if "资产负债" in sname:
                self._parse_balance_sheet(ws, stmts)
            elif "利润" in sname:
                self._parse_income_statement(ws, stmts)
            elif "现金流量" in sname:
                self._parse_cash_flow(ws, stmts)

        return stmts

    def _parse_balance_sheet(self, ws, stmts):
        """扫全表,左侧科目(col A)+右侧科目(col E),各取期末+年初。"""
        # 先嗅探报告日期和编制单位
        for i, row in enumerate(ws.iter_rows(values_only=True, max_row=6)):
            for cell in row:
                if cell is None:
                    continue
                s = str(cell)
                if "编制单位" in s:
                    m = re.search(r"编制单位[:：]?\s*(.+)", s)
                    if m:
                        stmts.reporting_entity = m.group(1).strip()
                if re.search(r"20\d{2}-\d{1,2}-\d{1,2}", s) and not stmts.statement_date:
                    m = re.search(r"(20\d{2}-\d{1,2}-\d{1,2})", s)
                    if m:
                        stmts.statement_date = m.group(1)

        # 逐行扫科目
        for row in ws.iter_rows(values_only=True, max_row=60):
            if not row or len(row) < 4:
                continue

            left_name = _norm(row[0]) if len(row) > 0 else ""
            left_cur = _parse_num(row[2]) if len(row) > 2 else None
            left_prev = _parse_num(row[3]) if len(row) > 3 else None

            right_name = _norm(row[4]) if len(row) > 4 else ""
            right_cur = _parse_num(row[6]) if len(row) > 6 else None
            right_prev = _parse_num(row[7]) if len(row) > 7 else None

            # 左侧资产
            for std in BS_LEFT_SUBJECTS:
                if _norm(std) == left_name:
                    stmts.balance_sheet[std] = PeriodValue(
                        current=left_cur, previous=left_prev, unit="万元",
                        current_label="2025末", previous_label="2024末",
                        compare_caption="较年初",
                    )
                    break

            # 右侧负债/权益(含同义项)
            for std in BS_RIGHT_SUBJECTS:
                if _norm(std) == right_name:
                    key = std
                    # 归一化几个别名
                    if std in ("实收资本(或股本)",):
                        key = "实收资本"
                    elif std in ("所有者权益(或股东权益)合计",):
                        key = "所有者权益合计"
                    elif std in ("负债和所有者权益(或股东权益)总计",):
                        key = "负债和所有者权益总计"
                    elif std == "库存股":
                        key = "库存股"
                    stmts.balance_sheet[key] = PeriodValue(
                        current=right_cur, previous=right_prev, unit="万元",
                        current_label="2025末", previous_label="2024末",
                        compare_caption="较年初",
                    )
                    break

    def _parse_income_statement(self, ws, stmts):
        for row in ws.iter_rows(values_only=True, max_row=40):
            if not row or len(row) < 3:
                continue
            name = _norm(row[0])
            # 利润表:col A 项目,col B 行数,col C 当期,col D 上期
            cur = _parse_num(row[2]) if len(row) > 2 else None
            prev = _parse_num(row[3]) if len(row) > 3 else None

            for std in IS_SUBJECTS:
                if _norm(std) == name:
                    stmts.income_statement[std] = PeriodValue(
                        current=cur, previous=prev, unit="万元",
                        current_label="2025", previous_label="2024"
                    )
                    break

    def _parse_cash_flow(self, ws, stmts):
        for row in ws.iter_rows(values_only=True, max_row=45):
            if not row or len(row) < 2:
                continue
            name = _norm(row[0])
            val = _parse_num(row[1]) if len(row) > 1 else None

            for std in CF_SUBJECTS:
                if _norm(std) == name:
                    stmts.cash_flow[std] = PeriodValue(
                        current=val, previous=None, unit="万元",
                        current_label="2025", previous_label=""
                    )
                    break

    # ------- 指标计算 -------

    def compute_indicators(self, stmts: FinancialStatements) -> FinancialIndicators:
        ind = FinancialIndicators(raw_statements=stmts)
        bs = stmts.balance_sheet
        is_ = stmts.income_statement
        cf = stmts.cash_flow

        def _bs(name):
            return bs.get(name, PeriodValue(unit="万元", current_label="2025末", previous_label="2024末"))

        def _is(name):
            return is_.get(name, PeriodValue(unit="万元", current_label="2025", previous_label="2024"))

        # 规模
        ind.total_assets = _bs("资产总计")
        ind.total_liabilities = _bs("负债合计")
        ind.total_equity = _bs("所有者权益合计")
        ind.revenue = _is("营业收入")
        ind.net_profit = _is("净利润")

        # 盈利
        rev = _is("营业收入")
        cost = _is("营业成本")
        op_profit = _is("营业利润")
        net = _is("净利润")

        ind.gross_margin = self._ratio_pv(
            (rev.current - cost.current) if rev.current and cost.current is not None else None,
            rev.current,
            (rev.previous - cost.previous) if rev.previous and cost.previous is not None else None,
            rev.previous,
            "%", rev.current_label, rev.previous_label,
        )
        ind.net_margin = self._ratio_pv(
            net.current, rev.current, net.previous, rev.previous,
            "%", rev.current_label, rev.previous_label,
        )
        ind.operating_margin = self._ratio_pv(
            op_profit.current, rev.current, op_profit.previous, rev.previous,
            "%", rev.current_label, rev.previous_label,
        )

        # 偿债 (BS 衍生指标,对比口径为"较年初")
        ind.debt_to_asset_ratio = self._ratio_pv(
            ind.total_liabilities.current, ind.total_assets.current,
            ind.total_liabilities.previous, ind.total_assets.previous,
            "%", "2025末", "2024末",
            compare_caption="较年初",
        )

        ca = _bs("流动资产合计")
        cl = _bs("流动负债合计")
        inv = _bs("存货")

        ind.current_ratio = self._ratio_pv(
            ca.current, cl.current, ca.previous, cl.previous,
            "倍", "2025末", "2024末", percent=False,
            compare_caption="较年初",
        )
        quick_cur = (ca.current - inv.current) if (ca.current is not None and inv.current is not None) else None
        quick_prev = (ca.previous - inv.previous) if (ca.previous is not None and inv.previous is not None) else None
        ind.quick_ratio = self._ratio_pv(
            quick_cur, cl.current, quick_prev, cl.previous,
            "倍", "2025末", "2024末", percent=False,
            compare_caption="较年初",
        )

        # 营运
        ar = _bs("应收账款")
        # 应收周转天数 = 应收/营收 × 365
        ar_days_cur = (ar.current / rev.current * 365) if (ar.current and rev.current) else None
        ar_days_prev = (ar.previous / rev.previous * 365) if (ar.previous and rev.previous) else None
        ind.ar_turnover_days = PeriodValue(
            current=ar_days_cur, previous=ar_days_prev, unit="天",
            current_label="2025", previous_label="2024",
        )

        inv_days_cur = (inv.current / cost.current * 365) if (inv.current and cost.current) else None
        inv_days_prev = (inv.previous / cost.previous * 365) if (inv.previous and cost.previous) else None
        ind.inventory_turnover_days = PeriodValue(
            current=inv_days_cur, previous=inv_days_prev, unit="天",
            current_label="2025", previous_label="2024",
        )

        # 现金流
        ocf = cf.get("经营活动产生的现金流量净额")
        icf = cf.get("投资活动产生的现金流量净额")
        fcf = cf.get("筹资活动产生的现金流量净额")
        cash_from_sales = cf.get("销售商品、提供劳务收到的现金")

        ind.operating_cf_net = ocf.current if ocf else None
        ind.investing_cf_net = icf.current if icf else None
        ind.financing_cf_net = fcf.current if fcf else None

        if ind.operating_cf_net is not None and net.current:
            ind.ocf_to_net_income = ind.operating_cf_net / net.current

        if cash_from_sales and cash_from_sales.current and rev.current:
            ind.cash_received_to_revenue = cash_from_sales.current / rev.current

        return ind

    @staticmethod
    def _ratio_pv(num_cur, den_cur, num_prev, den_prev, unit, cur_label, prev_label, percent=True, compare_caption="同比"):
        cur = None
        prev = None
        if num_cur is not None and den_cur not in (None, 0):
            cur = num_cur / den_cur * (100 if percent else 1)
        if num_prev is not None and den_prev not in (None, 0):
            prev = num_prev / den_prev * (100 if percent else 1)
        return PeriodValue(
            current=cur, previous=prev, unit=unit,
            current_label=cur_label, previous_label=prev_label,
            compare_caption=compare_caption,
        )

    # ------- 异常 & 趋势 -------

    def detect_anomalies(self, ind: FinancialIndicators) -> list:
        anomalies = []
        bs = ind.raw_statements.balance_sheet if ind.raw_statements else {}

        inv = bs.get("存货", PeriodValue())
        ar = bs.get("应收账款", PeriodValue())
        rev_yoy = ind.revenue.yoy_pct

        if inv.yoy_pct is not None and rev_yoy is not None:
            if inv.yoy_pct - rev_yoy > 30:
                anomalies.append(f"存货较年初{inv.yoy_pct:+.1f}%,远超营收同比增速{rev_yoy:+.1f}%,存货积压信号")

        if ar.yoy_pct is not None and rev_yoy is not None:
            if ar.yoy_pct - rev_yoy > 20:
                anomalies.append(f"应收账款较年初{ar.yoy_pct:+.1f}%,远超营收同比增速{rev_yoy:+.1f}%,应收激增")

        if ind.operating_cf_net is not None and ind.net_profit.current:
            if ind.operating_cf_net < 0 and ind.net_profit.current > 0:
                anomalies.append(f"净利润{ind.net_profit.current/10000:.1f}万为正但经营现金流净额为负({ind.operating_cf_net/10000:.1f}万),利润质量存疑")

        if ind.ocf_to_net_income is not None and ind.net_profit.current and ind.net_profit.current > 0:
            if ind.ocf_to_net_income < 0.5:
                anomalies.append(f"经营现金流/净利润={ind.ocf_to_net_income:.2f},低于0.5,经营现金流质量偏弱")

        if ind.cash_received_to_revenue is not None:
            if ind.cash_received_to_revenue < 0.8 or ind.cash_received_to_revenue > 1.3:
                anomalies.append(f"销售收现/营收={ind.cash_received_to_revenue:.2f},偏离正常区间[0.8-1.3]")

        short_debt = bs.get("短期借款", PeriodValue())
        if short_debt.current and ind.operating_cf_net and ind.operating_cf_net > 0:
            if short_debt.current > ind.operating_cf_net * 5:
                anomalies.append(f"短期借款{short_debt.current/10000:.0f}万,年经营现金流仅{ind.operating_cf_net/10000:.0f}万,短贷需滚动偿还")

        if ind.debt_to_asset_ratio.yoy_abs is not None and ind.debt_to_asset_ratio.yoy_abs > 10:
            anomalies.append(f"资产负债率较年初上升{ind.debt_to_asset_ratio.yoy_abs:+.1f}pct,杠杆显著上升")

        if ind.gross_margin.yoy_abs is not None and ind.gross_margin.yoy_abs < -5:
            anomalies.append(f"毛利率同比{ind.gross_margin.yoy_abs:+.1f}pct,显著下滑")

        return anomalies

    def summarize_trends(self, ind: FinancialIndicators) -> dict:
        labels = {}

        if ind.net_margin.yoy_abs is not None:
            if ind.net_margin.yoy_abs > 0.5:
                labels["盈利能力"] = "改善"
            elif ind.net_margin.yoy_abs < -0.5:
                labels["盈利能力"] = "恶化"
            else:
                labels["盈利能力"] = "稳定"

        if ind.debt_to_asset_ratio.yoy_abs is not None:
            if ind.debt_to_asset_ratio.yoy_abs < -2:
                labels["偿债"] = "改善"
            elif ind.debt_to_asset_ratio.yoy_abs > 2:
                labels["偿债"] = "恶化"
            else:
                labels["偿债"] = "稳定"

        if ind.ar_turnover_days.yoy_abs is not None:
            if ind.ar_turnover_days.yoy_abs < -10:
                labels["应收周转"] = "改善"
            elif ind.ar_turnover_days.yoy_abs > 10:
                labels["应收周转"] = "恶化"
            else:
                labels["应收周转"] = "稳定"

        if ind.inventory_turnover_days.yoy_abs is not None:
            if ind.inventory_turnover_days.yoy_abs < -10:
                labels["存货周转"] = "改善"
            elif ind.inventory_turnover_days.yoy_abs > 10:
                labels["存货周转"] = "恶化"
            else:
                labels["存货周转"] = "稳定"

        if ind.ocf_to_net_income is not None:
            if ind.ocf_to_net_income >= 1:
                labels["现金流质量"] = "健康"
            elif ind.ocf_to_net_income >= 0.5:
                labels["现金流质量"] = "偏弱"
            else:
                labels["现金流质量"] = "异常"

        if ind.revenue.yoy_pct is not None:
            if ind.revenue.yoy_pct > 10:
                labels["规模增长"] = "增长"
            elif ind.revenue.yoy_pct < -10:
                labels["规模增长"] = "收缩"
            else:
                labels["规模增长"] = "平稳"

        return labels

    # ------- 端到端 -------

    def analyze(self, file_path: str) -> FinancialIndicators:
        stmts = self.parse_xlsx(file_path)
        ind = self.compute_indicators(stmts)
        ind.anomalies = self.detect_anomalies(ind)
        ind.trend_labels = self.summarize_trends(ind)
        return ind

    # ------- 多年罗列块(V14 G1 新增) -------

    # 指标元数据:key 是 multi_year_data 输入中的 metric 名,value 描述展示方式
    #   display: prompt 中的中文标签
    #   unit: '万元' / '%' / '倍' / '天' (决定格式化方式)
    #   aliases: multi_year_data 里可能的别名(用户输入键名兼容)
    _MULTI_YEAR_METRIC_SPECS = [
        ("revenue", "营业收入", "万元",
         ["revenue", "operating_revenue", "营业收入", "主营业务收入"]),
        ("net_profit", "净利润", "万元",
         ["net_profit", "净利润"]),
        ("gross_margin", "毛利率", "%",
         ["gross_margin", "毛利率"]),
        ("net_margin", "净利率", "%",
         ["net_margin", "净利率"]),
        ("debt_to_asset_ratio", "资产负债率", "%",
         ["debt_to_asset_ratio", "资产负债率"]),
        ("current_ratio", "流动比率", "倍",
         ["current_ratio", "流动比率"]),
        ("ar_turnover_days", "应收账款周转天数", "天",
         ["ar_turnover_days", "应收账款周转天数", "应收周转天数"]),
        ("inventory_turnover_days", "存货周转天数", "天",
         ["inventory_turnover_days", "存货周转天数", "存货周转"]),
    ]

    @staticmethod
    def _fmt_metric_value(val, unit: str) -> str | None:
        """按单位格式化单期数值。返回 None 说明该期缺数。"""
        if val is None:
            return None
        try:
            v = float(str(val).replace(",", "").replace("%", ""))
        except (ValueError, TypeError):
            return None
        if unit == "万元":
            # 若传入值 >= 1e5 则按"元"处理折算万元;否则已是"万元"单位
            if abs(v) >= 1e5:
                v = v / 10000.0
            return f"{v:.1f}万"
        if unit == "%":
            return f"{v:.2f}%"
        if unit == "倍":
            return f"{v:.2f}倍"
        if unit == "天":
            return f"{v:.0f}天"
        return f"{v:.2f}"

    @staticmethod
    def _trend_phrase(values: list[tuple[str, float]], unit: str) -> str:
        """按 (年份, 数值) 序列判断趋势短语。代码判,不交给 LLM。

        返回如:"逐年提升"、"逐年下降"、"2024 触底后反弹"、"同比波动无明确方向"。
        values 已按年份升序。
        """
        if len(values) < 2:
            return ""
        labels = [v[0] for v in values]
        nums = [v[1] for v in values]

        # 单调判断
        diffs = [nums[i + 1] - nums[i] for i in range(len(nums) - 1)]
        all_up = all(d > 0 for d in diffs)
        all_down = all(d < 0 for d in diffs)
        if all_up:
            pct = (nums[-1] - nums[0]) / abs(nums[0]) * 100 if nums[0] else None
            base = "逐年提升" if unit in ("%", "倍", "天") else "逐年增长"
            if pct is not None:
                return f"{base},累计较{labels[0]}{'+' if pct>=0 else ''}{pct:.1f}%"
            return base
        if all_down:
            pct = (nums[-1] - nums[0]) / abs(nums[0]) * 100 if nums[0] else None
            base = "逐年下降"
            if pct is not None:
                return f"{base},累计较{labels[0]}{pct:.1f}%"
            return base

        # 谷底/顶点反弹:找最小值/最大值位置
        min_idx = nums.index(min(nums))
        max_idx = nums.index(max(nums))
        last_idx = len(nums) - 1
        # 先跌后升:最小值在中间且最后一期 > 最小值
        if 0 < min_idx < last_idx and nums[last_idx] > nums[min_idx]:
            recover_pct = (nums[last_idx] - nums[min_idx]) / abs(nums[min_idx]) * 100 if nums[min_idx] else None
            if recover_pct is not None and abs(recover_pct) >= 5:
                return f"{labels[min_idx]}触底后反弹,较谷底{'+' if recover_pct>=0 else ''}{recover_pct:.1f}%"
            return f"{labels[min_idx]}触底后反弹"
        # 先升后跌
        if 0 < max_idx < last_idx and nums[last_idx] < nums[max_idx]:
            return f"{labels[max_idx]}见顶后回落"

        # 末期同比方向
        if len(nums) >= 2:
            last_diff_pct = (nums[-1] - nums[-2]) / abs(nums[-2]) * 100 if nums[-2] else None
            if last_diff_pct is not None and abs(last_diff_pct) >= 5:
                direction = "回升" if last_diff_pct > 0 else "回落"
                return f"同比波动无明确方向,末期{labels[-1]}同比{direction}{abs(last_diff_pct):.1f}%"
        return "同比波动无明确方向"

    def _extract_metric_from_multiyear(self, multi_year_data: dict, aliases: list[str]) -> dict[str, float]:
        """从用户提供的 multi_year_data 按年份抽取某指标。返回 {year: float}。

        multi_year_data 结构:{年份: {metric_key: value}}。年份支持 '2022' 或 '2022.12' 等,
        此处按 re.fullmatch r"20\\d{2}" 过滤年度快照。
        """
        out: dict[str, float] = {}
        if not multi_year_data:
            return out
        for year_key, period_data in multi_year_data.items():
            if not isinstance(period_data, dict):
                continue
            ys = str(year_key).strip()
            if not re.fullmatch(r"20\d{2}", ys):
                continue
            for al in aliases:
                if al in period_data:
                    raw = period_data[al]
                    try:
                        out[ys] = float(str(raw).replace(",", "").replace("%", ""))
                    except (ValueError, TypeError):
                        continue
                    break
        return out

    def _format_multi_year_block(
        self,
        ind: FinancialIndicators,
        multi_year_data: dict | None,
    ) -> str:
        """生成【近三年财务指标罗列】块。当无 multi_year_data 时退化为 2 期展示。"""
        # 合并:优先以 multi_year_data 为真源;若未提供,降级只用 ind 的 current/previous
        # 目标:至少覆盖 8 项指标
        if not multi_year_data:
            # 降级:从 ind 的前/当期生成 2 期罗列(仍比无块好)
            fallback: dict[str, dict] = {}

            def _put(year_label: str, metric: str, val):
                if val is None:
                    return
                fallback.setdefault(year_label, {})[metric] = val

            # 年度标签从 ind.revenue 推断(如 '2025'/'2024')
            cur_lbl = (ind.revenue.current_label or "").strip()
            prev_lbl = (ind.revenue.previous_label or "").strip()
            cur_lbl = re.sub(r"末$", "", cur_lbl)
            prev_lbl = re.sub(r"末$", "", prev_lbl)
            # 还得是 20xx
            if not re.fullmatch(r"20\d{2}", cur_lbl) or not re.fullmatch(r"20\d{2}", prev_lbl):
                return ""

            _put(cur_lbl, "revenue", ind.revenue.current)
            _put(prev_lbl, "revenue", ind.revenue.previous)
            _put(cur_lbl, "net_profit", ind.net_profit.current)
            _put(prev_lbl, "net_profit", ind.net_profit.previous)
            _put(cur_lbl, "gross_margin", ind.gross_margin.current)
            _put(prev_lbl, "gross_margin", ind.gross_margin.previous)
            _put(cur_lbl, "net_margin", ind.net_margin.current)
            _put(prev_lbl, "net_margin", ind.net_margin.previous)
            _put(cur_lbl, "debt_to_asset_ratio", ind.debt_to_asset_ratio.current)
            _put(prev_lbl, "debt_to_asset_ratio", ind.debt_to_asset_ratio.previous)
            _put(cur_lbl, "current_ratio", ind.current_ratio.current)
            _put(prev_lbl, "current_ratio", ind.current_ratio.previous)
            _put(cur_lbl, "ar_turnover_days", ind.ar_turnover_days.current)
            _put(prev_lbl, "ar_turnover_days", ind.ar_turnover_days.previous)
            _put(cur_lbl, "inventory_turnover_days", ind.inventory_turnover_days.current)
            _put(prev_lbl, "inventory_turnover_days", ind.inventory_turnover_days.previous)
            multi_year_data = fallback

        # 收集年度列表(升序)
        years: list[str] = sorted(
            [str(y) for y in multi_year_data.keys() if re.fullmatch(r"20\d{2}", str(y))]
        )
        if len(years) < 2:
            return ""

        lines: list[str] = []
        lines.append("【近三年财务指标罗列】(撰写财务分析段时必须原文引用,禁止省略中间年份)")
        lines.append(f"覆盖年度: {' / '.join(years)} (共{len(years)}年)")

        any_row = False
        for metric_key, display, unit, aliases in self._MULTI_YEAR_METRIC_SPECS:
            series = self._extract_metric_from_multiyear(multi_year_data, aliases)
            if not series:
                continue
            # 按年份序列化
            pairs: list[tuple[str, float]] = []
            for y in years:
                if y in series:
                    pairs.append((y, series[y]))
            if len(pairs) < 2:
                continue
            # 原始值字符串
            parts: list[str] = []
            for y, v in pairs:
                disp = self._fmt_metric_value(v, unit)
                if disp is None:
                    continue
                parts.append(f"{y}={disp}")
            if len(parts) < 2:
                continue
            trend = self._trend_phrase(pairs, unit)
            # 再补一句"分别为"样板,LLM 原文引用即可成三段式数据
            wording = "分别为 " + " 、 ".join(disp for _, disp in
                                              [(y, self._fmt_metric_value(v, unit)) for y, v in pairs]
                                              if disp)
            line = f"  · {display}: {' → '.join(parts)} ({wording};{trend})" if trend else \
                   f"  · {display}: {' → '.join(parts)} ({wording})"
            lines.append(line)
            any_row = True

        if not any_row:
            return ""

        lines.append("(口径:借款人单体报表;撰写时须按『分别为…』句式原文罗列年度数值,"
                     "再跟一句趋势短语,最后归因。禁止只写末期数或跳年。)")
        return "\n".join(lines)

    # ------- Prompt 注入文本 -------

    def format_for_prompt(
        self,
        ind: FinancialIndicators,
        multi_year_data: dict | None = None,
    ) -> str:
        """格式化为 prompt 注入文本。

        multi_year_data: 可选,三年/多年财务数据快照,格式为
            {年份(str): {metric_key(str): 数值或字符串}}
            metric_key 支持中英双语常见命名,本方法内部会做别名归一。
            当提供时,输出顶部追加【近三年财务指标罗列】块,供财务段"分别为"式
            罗列;LLM 被要求原文引用,禁止省略中间年份。
        """
        stmts = ind.raw_statements
        entity = stmts.reporting_entity if stmts and stmts.reporting_entity else "借款人"
        lines = []

        # 新增:多年罗列块(置顶,G2 生成层必须原文引用)
        multi_block = self._format_multi_year_block(ind, multi_year_data)
        if multi_block:
            lines.append(multi_block)
            lines.append("")

        lines.append("【已计算财务指标(代码从xlsx直接抽取并计算,数字权威,禁止重新计算)】")
        lines.append(f"口径说明:以下数据口径为{entity}单体报表(非合并),BS科目对比用『较年初』,IS/CF科目对比用『同比』")
        if stmts:
            lines.append(f"数据来源:{stmts.source_file}")
            if stmts.statement_date:
                lines.append(f"报告日:{stmts.statement_date}")
        lines.append("")

        lines.append("[规模]")
        lines.append(f"总资产: {ind.total_assets.format()}")
        lines.append(f"总负债: {ind.total_liabilities.format()}")
        lines.append(f"所有者权益: {ind.total_equity.format()}")
        lines.append(f"营业收入: {ind.revenue.format()}")
        lines.append(f"净利润: {ind.net_profit.format()}")
        lines.append("")

        lines.append("[盈利能力]")
        lines.append(f"毛利率: {ind.gross_margin.format()}")
        lines.append(f"净利率: {ind.net_margin.format()}")
        lines.append(f"营业利润率: {ind.operating_margin.format()}")
        lines.append("")

        lines.append("[偿债能力]")
        lines.append(f"资产负债率: {ind.debt_to_asset_ratio.format()}")
        lines.append(f"流动比率: {ind.current_ratio.format()}")
        lines.append(f"速动比率: {ind.quick_ratio.format()}")
        lines.append("")

        lines.append("[营运效率]")
        lines.append(f"应收账款周转天数: {ind.ar_turnover_days.format()}")
        lines.append(f"存货周转天数: {ind.inventory_turnover_days.format()}")
        lines.append("")

        lines.append("[现金流]")
        if ind.operating_cf_net is not None:
            lines.append(f"经营活动净现金流(当期): {ind.operating_cf_net/10000:.1f}万元")
        if ind.investing_cf_net is not None:
            lines.append(f"投资活动净现金流(当期): {ind.investing_cf_net/10000:.1f}万元")
        if ind.financing_cf_net is not None:
            lines.append(f"筹资活动净现金流(当期): {ind.financing_cf_net/10000:.1f}万元")
        if ind.ocf_to_net_income is not None:
            lines.append(f"经营现金流/净利润: {ind.ocf_to_net_income:.2f} (健康水平>=1)")
        if ind.cash_received_to_revenue is not None:
            lines.append(f"销售商品收现/营收: {ind.cash_received_to_revenue:.2f} (正常区间[0.8-1.3])")
        lines.append("")

        # 关键科目原始数据(供 LLM 引用)
        if stmts:
            lines.append("[关键科目原始期末值(万元,供引用)]")
            for subj in ["货币资金", "应收账款", "其他应收款", "存货",
                        "流动资产合计", "非流动资产合计",
                        "短期借款", "应付账款", "流动负债合计",
                        "实收资本"]:
                pv = stmts.balance_sheet.get(subj)
                if pv and pv.current is not None:
                    lines.append(f"  {subj}: {pv.format()}")
            lines.append("")

        if ind.anomalies:
            lines.append("[代码识别的异常项 / 需重点关注]")
            for a in ind.anomalies:
                lines.append(f"- {a}")
            lines.append("")

        if ind.trend_labels:
            lines.append("[趋势定性]")
            for k, v in ind.trend_labels.items():
                lines.append(f"{k}: {v}")
            lines.append("")

        # 行业基准参照(供 LLM 做归因对比,非硬指标)
        try:
            from industry_benchmark import get_benchmark
            bm = get_benchmark("软件信息")
            if bm is not None:
                lines.append(bm.format_for_prompt())
        except Exception:
            pass

        return "\n".join(lines)

    def get_subject_anchor_rules(self, ind: FinancialIndicators) -> str:
        """Phase2 专用:引用财务数字时的主体归属硬规则。
        不在 Phase1 证据组装中使用,避免 LLM 对非财务字段过度保守。"""
        stmts = ind.raw_statements
        entity = stmts.reporting_entity if stmts and stmts.reporting_entity else "借款人"
        return (
            f"【财务主体归属硬规则(仅约束财务数字的引用)】\n"
            f"1. 上文【已计算财务指标】块中的数字,仅代表借款人{entity}单体口径,"
            f"不得将其套用到母公司/合并报表/保证人/抵押人/关联企业/子公司/集团。\n"
            f"2. 禁止给{entity}贴『母公司』『控股公司』『集团』等角色标签,它就是借款人本身。\n"
            f"3. 该规则只约束财务数字引用;非财务字段(成立日期/地址/法定代表人/主营业务等)"
            f"按证据清单照常撰写,不受此规则影响。"
        )


# ============================================================
# 自测
# ============================================================

if __name__ == "__main__":
    import sys, io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', line_buffering=True)

    fp = r"D:\刘野\众安\新建文件夹\2026.3.25续贷材料\2、2025年12月财务报表-中锐20260212.xlsx"

    analyzer = FinancialAnalyzer()
    ind = analyzer.analyze(fp)
    print(analyzer.format_for_prompt(ind))

    print()
    print("=" * 60)
    print("关键数字交叉验证")
    print("=" * 60)

    def chk(label, got, expected, tol=0.5):
        if got is None:
            mark = "X 无"
        else:
            diff = abs(got - expected)
            mark = "OK" if diff <= tol else f"X差{diff:.2f}"
        print(f"{mark:>6}  {label}: got={got}, expected={expected}")

    chk("营收 YoY %", ind.revenue.yoy_pct, 14.9)
    ar = ind.raw_statements.balance_sheet.get("应收账款")
    chk("应收 YoY %", ar.yoy_pct if ar else None, -19.7)
    inv = ind.raw_statements.balance_sheet.get("存货")
    chk("存货 YoY %", inv.yoy_pct if inv else None, 88.0, tol=1)
    chk("资产负债率 当期 %", ind.debt_to_asset_ratio.current, 42.5, tol=0.5)
    chk("经营现金流(万元)", ind.operating_cf_net/10000 if ind.operating_cf_net else None, 791.2, tol=0.5)
    chk("总资产 当期(万元)", ind.total_assets.current/10000 if ind.total_assets.current else None, 16956.5, tol=1)
    chk("净利润 YoY %", ind.net_profit.yoy_pct, 50.1, tol=1)
