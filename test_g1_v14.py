# -*- coding: utf-8 -*-
"""test_g1_v14.py — V14 G1(数据层根基)单元测试。

验收:5/5 全绿。
  1. FinancialAnalyzer.format_for_prompt 注入【近三年财务指标罗列】块 + 趋势短语
  2. industry_cards.find_cards 命中/空两种场景
  3. MaterialAnchor.format_for_prompt 含【行业/政策参考卡片】块
  4. truth_fill.should_skip_consolidated_financial_table 无合并源时 skip + pending_tag 正确
  5. truth_fill.should_skip_consolidated_financial_table 有合并源时放行
"""

from __future__ import annotations

import os
import sys
import tempfile
import traceback


HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, HERE)


RESULTS: list[tuple[str, bool, str]] = []


def case(name: str):
    def deco(fn):
        try:
            fn()
            RESULTS.append((name, True, ""))
            print(f"[OK] {name}")
        except AssertionError as ae:
            RESULTS.append((name, False, str(ae)))
            print(f"[FAIL] {name}: {ae}")
        except Exception as e:
            tb = traceback.format_exc(limit=3)
            RESULTS.append((name, False, f"{e}\n{tb}"))
            print(f"[ERROR] {name}: {e}\n{tb}")
        return fn
    return deco


# =====================================================================
# Case 1: FinancialAnalyzer multi-year 罗列块
# =====================================================================

@case("1. FinancialAnalyzer.format_for_prompt 含【近三年财务指标罗列】块")
def _c1():
    from financial_analyzer import (
        FinancialAnalyzer,
        FinancialIndicators,
        PeriodValue,
        FinancialStatements,
    )

    # 构造 ind(2 期够跑降级逻辑,但我们给 multi_year_data 用 4 年真样本)
    stmts = FinancialStatements(source_file="<test>", reporting_entity="中锐网络科技")
    ind = FinancialIndicators(raw_statements=stmts)
    ind.revenue = PeriodValue(
        current=10010.0 * 10000, previous=8712.4 * 10000,
        unit="万元", current_label="2025", previous_label="2024",
    )
    ind.net_profit = PeriodValue(
        current=494.0 * 10000, previous=329.0 * 10000,
        unit="万元", current_label="2025", previous_label="2024",
    )

    multi_year_data = {
        "2022": {
            "revenue": 21198.0 * 10000,
            "net_profit": 983.0 * 10000,
            "gross_margin": 30.55,
            "net_margin": 4.64,
            "debt_to_asset_ratio": 55.0,
            "current_ratio": 1.20,
            "ar_turnover_days": 200,
            "inventory_turnover_days": 45,
        },
        "2023": {
            "revenue": 12130.5 * 10000,
            "net_profit": 393.0 * 10000,
            "gross_margin": 27.30,
            "net_margin": 3.24,
            "debt_to_asset_ratio": 50.0,
            "current_ratio": 1.25,
            "ar_turnover_days": 260,
            "inventory_turnover_days": 55,
        },
        "2024": {
            "revenue": 8712.4 * 10000,
            "net_profit": 329.0 * 10000,
            "gross_margin": 38.80,
            "net_margin": 3.78,
            "debt_to_asset_ratio": 45.0,
            "current_ratio": 1.35,
            "ar_turnover_days": 271,
            "inventory_turnover_days": 60,
        },
        "2025": {
            "revenue": 10010.0 * 10000,
            "net_profit": 494.0 * 10000,
            "gross_margin": 38.00,
            "net_margin": 4.94,
            "debt_to_asset_ratio": 42.5,
            "current_ratio": 1.40,
            "ar_turnover_days": 220,
            "inventory_turnover_days": 50,
        },
    }

    out = FinancialAnalyzer().format_for_prompt(ind, multi_year_data=multi_year_data)
    # 1.1 含罗列块标题
    assert "【近三年财务指标罗列】" in out, "罗列块标题缺失"
    # 1.2 至少 4 年全部出现在输出里
    for y in ("2022", "2023", "2024", "2025"):
        assert y + "=" in out, f"{y}= 格式缺失"
    # 1.3 至少 6 项指标出现(覆盖 >=8 的指标中的大部分)
    for metric in ["营业收入", "净利润", "毛利率", "净利率",
                   "资产负债率", "流动比率",
                   "应收账款周转天数", "存货周转天数"]:
        assert metric in out, f"指标 {metric} 未出现"
    # 1.4 存在"分别为" pattern
    assert "分别为" in out, "缺 『分别为』 句式,LLM 无法原文引用"
    # 1.5 趋势短语:营业收入 2022-2024 下降、2025 反弹,应出现"触底"或"反弹"
    assert ("触底" in out) or ("反弹" in out), "营业收入趋势短语未命中谷底反弹"

    print("  multi-year block preview:")
    for line in out.splitlines()[:14]:
        print("   " + line)


# =====================================================================
# Case 2: industry_cards.find_cards 匹配/空
# =====================================================================

@case("2. industry_cards.find_cards 命中 / 空")
def _c2():
    from industry_cards import find_cards

    r1 = find_cards(["软件", "信息技术"])
    assert r1, "软件+信息技术 无命中,期望返回 软件和信息技术服务业 卡"
    assert any(c.industry == "软件和信息技术服务业" for c in r1), \
        f"top_k 未包含软件卡,got={[c.industry for c in r1]}"

    r2 = find_cards(["水利"])
    assert any(c.industry == "智慧水利" for c in r2), \
        f"水利关键词未命中智慧水利,got={[c.industry for c in r2]}"

    r3 = find_cards(["unknown_xyz_industry"])
    assert r3 == [], f"未知关键词应返回空列表,got={r3}"

    # 验证 schema 完整性
    for c in r1:
        assert c.industry and c.benchmarks and c.source, \
            f"卡片字段缺失: {c.to_dict()}"


# =====================================================================
# Case 3: MaterialAnchor.format_for_prompt 含行业卡片块
# =====================================================================

@case("3. MaterialAnchor.format_for_prompt 含【行业/政策参考卡片】块")
def _c3():
    from material_anchor import build_material_anchor

    # 无 file_path_map,通过 extra_keywords 触发匹配
    anchor = build_material_anchor(
        file_path_map={},
        extra_keywords=["软件开发", "信息技术服务"],
        top_k=2,
    )
    text = anchor.format_for_prompt()
    assert "【行业/政策参考卡片】" in text, \
        f"卡片块标题缺失;format 输出前 200 字符:{text[:200]!r}"
    assert "软件和信息技术服务业" in text or "卡片: 软件" in text, \
        "软件卡片 industry 行未出现"
    # 政策/趋势/来源均应出现
    assert "政策:" in text, "政策行缺失"
    assert "趋势:" in text, "趋势行缺失"
    assert "来源:" in text, "来源行缺失"


# =====================================================================
# Case 4: 无合并报表源 -> skip + pending_tag
# =====================================================================

@case("4. truth_fill.should_skip_consolidated_financial_table 无合并源时 skip")
def _c4():
    from truth_fill import should_skip_consolidated_financial_table

    # 构造一个仅含借款人单体报表的 xlsx(用 openpyxl 写,sheet 名全是"单体")
    from openpyxl import Workbook
    tmpdir = tempfile.mkdtemp(prefix="g1test_")
    xlsx_path = os.path.join(tmpdir, "单体财务报表.xlsx")
    wb = Workbook()
    wb.active.title = "资产负债表"
    wb.create_sheet("利润表")
    wb.create_sheet("现金流量表")
    wb.save(xlsx_path)

    file_path_map = {"单体财务报表.xlsx": xlsx_path}
    header = "2.集团合并财务数据(若有):"
    should_skip, tag = should_skip_consolidated_financial_table(
        header, file_path_map, material_anchor=None
    )
    assert should_skip is True, \
        f"无合并报表源但未 skip,got should_skip={should_skip}, tag={tag}"
    assert "合并财务报表" in tag, \
        f"pending_tag 不含『合并财务报表』,got={tag}"

    # 非"合并"类表 header,不应 skip
    s2, t2 = should_skip_consolidated_financial_table(
        "1.借款人基本财务数据:", file_path_map, material_anchor=None
    )
    assert s2 is False and t2 == "", \
        f"非合并表被误 skip,got=({s2},{t2})"


# =====================================================================
# Case 5: 有合并报表源 -> 放行
# =====================================================================

@case("5. truth_fill.should_skip_consolidated_financial_table 有合并源时放行")
def _c5():
    from truth_fill import should_skip_consolidated_financial_table

    from openpyxl import Workbook
    tmpdir = tempfile.mkdtemp(prefix="g1test_")
    xlsx_path = os.path.join(tmpdir, "集团合并报表.xlsx")
    wb = Workbook()
    wb.active.title = "合并资产负债表"
    wb.create_sheet("合并利润表")
    wb.create_sheet("合并现金流量表")
    wb.save(xlsx_path)

    file_path_map = {"集团合并报表.xlsx": xlsx_path}
    should_skip, tag = should_skip_consolidated_financial_table(
        "2.集团合并财务数据(若有):", file_path_map, material_anchor=None
    )
    assert should_skip is False, \
        f"有合并 sheet 仍被 skip,got should_skip={should_skip}, tag={tag}"


# =====================================================================
if __name__ == "__main__":
    total = len(RESULTS)
    ok = sum(1 for _, p, _ in RESULTS if p)
    print("")
    print("=" * 60)
    print(f"Result: {ok}/{total} passed")
    for name, passed, err in RESULTS:
        mark = "PASS" if passed else "FAIL"
        print(f"  [{mark}] {name}")
        if not passed and err:
            first_line = err.splitlines()[0] if err else ""
            print(f"         {first_line}")
    sys.exit(0 if ok == total else 1)
