import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', line_buffering=True)

from openpyxl import load_workbook

fp = r"D:\刘野\众安\新建文件夹\2026.3.25续贷材料\2、2025年12月财务报表-中锐20260212.xlsx"
wb = load_workbook(fp, data_only=True)
print("SHEETS:", wb.sheetnames)

for name in wb.sheetnames:
    ws = wb[name]
    print(f"\n===== {name} (dim={ws.dimensions}, max_row={ws.max_row}, max_col={ws.max_column}) =====")
    for r in range(1, ws.max_row + 1):
        row_vals = []
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if v is None:
                row_vals.append("")
            else:
                row_vals.append(str(v))
        # 只打印非全空行
        if any(x.strip() for x in row_vals):
            print(f"R{r}: " + " | ".join(f"[{chr(64+i+1)}]{v}" for i, v in enumerate(row_vals) if v))
