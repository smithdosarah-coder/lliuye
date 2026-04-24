# -*- coding: utf-8 -*-
"""Agent4 — 在贷榜单专用导出器

基于 shared.kb_scan.exporters 做增强：
- 按 🔴/🟡/🟢 分 3 个 sheet
- 列扩展：授信额度 / 贷款余额 / 贷款到期日 / 客户经理 / 内部评级 / 外部命中 / 内部命中 / 处置建议
- 顶部汇总行 + 分级行染色
"""

from __future__ import annotations
from pathlib import Path
from datetime import datetime

from shared.kb_scan.models import HitList, RiskLevel


LEVEL_META = {
    RiskLevel.RED:    {"label": "🔴 高风险", "fill": "FFE2E2", "order": 0},
    RiskLevel.YELLOW: {"label": "🟡 中风险", "fill": "FFF6CC", "order": 1},
    RiskLevel.GREEN:  {"label": "🟢 低风险", "fill": "E8F6E2", "order": 2},
}


def export_ledger_excel(
    hl: HitList,
    dispositions: dict[str, "DispositionPlan"] | None = None,  # noqa: F821
    output_dir: str | Path = "outputs",
) -> str:
    """导出在贷榜单 Excel。

    Args:
        hl: HitList
        dispositions: company_name -> DispositionPlan，可选
    """
    out_dir = Path(output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = out_dir / f"alert_ledger_{ts}.xlsx"

    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment
    except (ImportError, ModuleNotFoundError, AttributeError):
        return _export_csv(hl, out_dir, ts)

    wb = Workbook()
    wb.remove(wb.active)

    headers = [
        "排名", "企业名称", "授信额度(万)", "贷款余额(万)", "贷款到期日",
        "客户经理", "内评", "外部命中", "内部命中", "命中规则", "核心理由", "处置建议",
    ]

    groups = [
        ("高风险_红灯", [h for h in hl.hits if h.level == RiskLevel.RED]),
        ("中风险_黄灯", [h for h in hl.hits if h.level == RiskLevel.YELLOW]),
        ("低风险_绿灯", [h for h in hl.hits if h.level == RiskLevel.GREEN]),
    ]

    for sheet_name, items in groups:
        ws = wb.create_sheet(title=sheet_name)
        fill_color = LEVEL_META.get(
            items[0].level if items else RiskLevel.GREEN, {}
        ).get("fill", "FFFFFF")
        # 顶部汇总
        label = LEVEL_META[items[0].level]["label"] if items else sheet_name
        ws.append([f"{label} — 共 {len(items)} 家"])
        ws["A1"].font = Font(bold=True, size=13)
        ws.append([f"生成时间 {hl.generated_at} · 知识库 {hl.kb_summary}"])
        ws.append([hl.scan_summary])
        ws.append([])
        ws.append(headers)
        header_row = ws.max_row
        for c in ws[header_row]:
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal="center", vertical="center")

        for h in items:
            payload = h.target.payload
            extras = payload.get("extras", {}) or {}
            ext_hits = h.extras.get("external_hits", []) if h.extras else []
            int_hits = h.extras.get("internal_hits", []) if h.extras else []
            disposition_text = ""
            plan = (dispositions or {}).get(payload.get("company_name", ""))
            if plan is not None:
                disposition_text = "\n".join(
                    f"[{a.urgency}] {a.description[:24]} ({a.responsible})"
                    for a in plan.actions[:3]
                )
            ws.append([
                h.rank,
                payload.get("company_name", ""),
                extras.get("credit_line", ""),
                extras.get("outstanding", ""),
                extras.get("due_date", ""),
                extras.get("manager", ""),
                extras.get("internal_rating", ""),
                ", ".join(ext_hits) or "-",
                ", ".join(int_hits) or "-",
                ", ".join(h.matched_rules) or "-",
                " | ".join(h.reasons[:3]) or "-",
                disposition_text or "-",
            ])
            for c in ws[ws.max_row]:
                c.fill = PatternFill(start_color=fill_color,
                                     end_color=fill_color, fill_type="solid")
                c.alignment = Alignment(wrap_text=True, vertical="top")

        widths = [6, 28, 12, 12, 12, 10, 6, 18, 18, 24, 40, 36]
        for i, w in enumerate(widths):
            ws.column_dimensions[chr(ord("A") + i)].width = w

    # 汇总 sheet
    summary_ws = wb.create_sheet(title="汇总", index=0)
    summary_ws.append(["贷中风险预警雷达 — 扫描汇总"])
    summary_ws["A1"].font = Font(bold=True, size=15)
    summary_ws.append([f"生成时间：{hl.generated_at}"])
    summary_ws.append([f"知识库：{hl.kb_summary}"])
    summary_ws.append([f"扫描范围：{hl.scan_summary}"])
    summary_ws.append([])
    summary_ws.append(["分级", "家数"])
    summary_ws.append(["🔴 高风险", hl.red_count])
    summary_ws.append(["🟡 中风险", hl.yellow_count])
    summary_ws.append(["🟢 低风险", hl.green_count])
    summary_ws.append(["合计", hl.total_scanned])
    summary_ws.column_dimensions["A"].width = 30
    summary_ws.column_dimensions["B"].width = 14

    wb.save(str(path))
    return str(path)


def _export_csv(hl: HitList, out_dir: Path, ts: str) -> str:
    import csv
    path = out_dir / f"alert_ledger_{ts}.csv"
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        w.writerow([
            "分级", "排名", "企业名称", "授信额度(万)", "贷款余额(万)",
            "贷款到期日", "客户经理", "命中规则", "核心理由",
        ])
        for h in hl.hits:
            payload = h.target.payload
            extras = payload.get("extras", {}) or {}
            w.writerow([
                h.level.value, h.rank, payload.get("company_name", ""),
                extras.get("credit_line", ""), extras.get("outstanding", ""),
                extras.get("due_date", ""), extras.get("manager", ""),
                ", ".join(h.matched_rules), " | ".join(h.reasons[:3]),
            ])
    return str(path)
