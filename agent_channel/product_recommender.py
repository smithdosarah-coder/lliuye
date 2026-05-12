# -*- coding: utf-8 -*-
"""ProductRecommender: 对每条线索调 channel_rules + scoring 推荐 Top3 产品，并生成切入话术。

完全复用 v1.0 的 channel_rules.match_channels() 和 scoring.rank_recommendations()，
作为"产品推荐器"二次封装。LLM 只负责"话术"这一层，失败可降级到模板。
"""
from __future__ import annotations

import json
import os
import re
from pathlib import Path
from typing import Iterable

from shared.kb_scan.models import HitItem, IdealProfile
from agent_channel.channel_rules import match_channels
from agent_channel.scoring import rank_recommendations, ChannelRecommendation
from shared.prompts.agent_helpers import build_channel_ssot_prompt
from agent_channel.prompts import (
    PITCH_GEN_PROMPT,
    BATCH_PITCH_PROMPT,
    render_fewshot_block,
)


def load_product_catalog(kb_dir: str | Path) -> list[dict]:
    """装载 data/mock/channel-kb/product-catalog/*.xlsx 为 row dict 列表。

    返回字段约定:
      product / positioning / segment / amount_range /
      term / rate / guarantee / risk_gates
    openpyxl 缺失或空目录返回 []。
    """
    base = Path(kb_dir)
    if not base.is_dir():
        return []
    try:
        import openpyxl
    except ImportError:
        return []
    rows: list[dict] = []
    keys = ("product", "positioning", "segment", "amount_range",
            "term", "rate", "guarantee", "risk_gates")
    for fname in sorted(os.listdir(base)):
        if not fname.lower().endswith(".xlsx"):
            continue
        fpath = base / fname
        try:
            wb = openpyxl.load_workbook(str(fpath), data_only=True)
        except (OSError, ValueError, KeyError):
            continue
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            data = list(ws.iter_rows(values_only=True))
            if not data or len(data) < 2:
                continue
            header = data[0]
            # 跳过"说明" 类非产品 sheet
            if not header or not any(
                isinstance(v, str) and any(k in v for k in ("产品", "名称"))
                for v in header if v
            ):
                continue
            for row in data[1:]:
                if not row or all(v is None or str(v).strip() == "" for v in row):
                    continue
                mapping = dict(zip(keys, [str(v).strip() if v is not None else "" for v in row[:len(keys)]]))
                if not mapping.get("product"):
                    continue
                mapping["source_doc"] = str(fpath)
                mapping["source_sheet"] = sheet_name
                rows.append(mapping)
    return rows


class ProductRecommender:
    """产品推荐 + 话术生成。"""

    def __init__(self, llm_chat=None, llm_json=None):
        """llm_chat: (system, user) -> str；llm_json: (system, user) -> dict。均可为 None。"""
        self.llm_chat = llm_chat
        self.llm_json = llm_json

    # ---- 主入口：产品匹配（确定性）----

    def recommend_products(self, hit: HitItem, top_n: int = 3) -> list[ChannelRecommendation]:
        company_info = self._hit_to_company_info(hit)
        matched = match_channels(company_info)
        if not matched:
            # 给一个保底"通用方案"
            matched = [{
                "category": "综合推荐",
                "channel_name": "小微企业信用贷",
                "max_amount": "1000万",
                "match_reasons": ["基于企业基础信息推荐"],
                "dimensions_met": 1,
            }]
        recs = rank_recommendations(company_info, matched)
        return recs[:top_n]

    def enrich_hit(self, hit: HitItem, ideal: IdealProfile, top_n: int = 3) -> None:
        """把 Top3 产品注入 hit.extras，话术留空（由 batch_generate_pitches 批量灌入）。"""
        top = self.recommend_products(hit, top_n=top_n)
        hit.extras["recommended_products"] = [r.model_dump() for r in top]
        hit.extras.setdefault("pitch_script", "")

    # ---- 批量话术（推荐入口，降低 LLM 调用次数）----

    def batch_generate_pitches(
        self,
        hits: Iterable[HitItem],
        ideal: IdealProfile,
    ) -> dict[str, str]:
        """对一组已经 enrich 过的 hit 批量生成话术，返回 {hit_id: pitch}。"""
        hit_list = list(hits)
        if not hit_list:
            return {}

        # 无 LLM → 模板兜底
        if self.llm_json is None:
            return {h.hit_id: self._template_pitch(h, ideal) for h in hit_list}

        leads_block_lines = []
        for h in hit_list:
            products = h.extras.get("recommended_products") or []
            payload = h.target.payload
            products_str = " / ".join(
                f"{p.get('channel_name','')}(额度{p.get('max_amount','')})"
                for p in products[:3]
            ) or "（暂无产品）"
            leads_block_lines.append(
                "- lead_id: {hid}\n"
                "  企业: {name} / {industry} / {scale} / 营收{rev}\n"
                "  标签: {tags}\n"
                "  推荐产品: {prods}".format(
                    hid=h.hit_id,
                    name=payload.get("company_name", ""),
                    industry=payload.get("industry", "") or "未知",
                    scale=payload.get("scale", "") or "未知",
                    rev=payload.get("revenue_latest", "") or "未知",
                    tags="/".join((payload.get("tags") or [])[:3]) or "无",
                    prods=products_str,
                )
            )
        leads_block = "\n".join(leads_block_lines)

        try:
            system = build_channel_ssot_prompt(
                task_type="pitch_batch",
                schema_hint='{"pitches": [{"lead_id": str, "pitch": str (80-120 字 · 口语化 · 无客套)}]}',
            )
            data = self.llm_json(
                system,
                BATCH_PITCH_PROMPT.format(
                    policy_ctx=ideal.policy_context or "（无政策语境）",
                    leads_block=leads_block,
                ),
            )
        except (RuntimeError, ValueError, TypeError, OSError, AttributeError, KeyError, ImportError):
            data = None

        result: dict[str, str] = {}
        if isinstance(data, dict) and isinstance(data.get("pitches"), list):
            for p in data["pitches"]:
                if isinstance(p, dict):
                    hid = p.get("lead_id") or ""
                    pitch = (p.get("pitch") or "").strip()
                    if hid and pitch:
                        result[hid] = pitch

        # 缺失的用模板兜底
        for h in hit_list:
            if h.hit_id not in result:
                result[h.hit_id] = self._template_pitch(h, ideal)

        return result

    # ---- 单条话术（fallback 路径）----

    def generate_pitch_single(self, hit: HitItem, ideal: IdealProfile) -> str:
        if self.llm_chat is None:
            return self._template_pitch(hit, ideal)
        payload = hit.target.payload
        products = hit.extras.get("recommended_products") or []
        products_str = "\n".join(
            f"- {p.get('channel_name','')}（{p.get('category','')}，额度{p.get('max_amount','')}，{p.get('score','')}分）"
            for p in products[:3]
        ) or "- 小微企业信用贷（普惠金融，额度1000万）"
        user = PITCH_GEN_PROMPT.format(
            company_name=payload.get("company_name", ""),
            industry=payload.get("industry", ""),
            sub_industry=payload.get("sub_industry", ""),
            scale=payload.get("scale", ""),
            revenue=payload.get("revenue_latest", ""),
            main_business=payload.get("main_business", ""),
            tags="/".join((payload.get("tags") or [])[:5]),
            products=products_str,
            policy_ctx=ideal.policy_context or "（无政策语境）",
            fewshot_block=render_fewshot_block(),
        )
        try:
            system = build_channel_ssot_prompt(task_type="pitch_single")
            out = self.llm_chat(system, user)
            return (out or "").strip().strip('"""\'') or self._template_pitch(hit, ideal)
        except (RuntimeError, ValueError, TypeError, OSError, AttributeError, KeyError, ImportError):
            return self._template_pitch(hit, ideal)

    # ---- 内部：Hit payload → match_channels 输入 ----

    @staticmethod
    def _hit_to_company_info(hit: HitItem) -> dict:
        p = hit.target.payload or {}
        keywords = p.get("keywords") or []
        if isinstance(keywords, list):
            keywords_str = " ".join(keywords)
        else:
            keywords_str = str(keywords)
        tags = p.get("tags") or []
        if isinstance(tags, list):
            keywords_str = keywords_str + " " + " ".join(tags)
        quals = p.get("qualifications") or []
        if isinstance(quals, list):
            keywords_str = keywords_str + " " + " ".join(quals)
        return {
            "company_name": p.get("company_name", ""),
            "industry": p.get("industry", ""),
            "region": p.get("region", ""),
            "scale": p.get("scale", ""),
            "keywords": keywords_str.strip(),
        }

    # ---- Batch 2 · catalog 反查(走规则,不走 LLM) ----

    @staticmethod
    def recommend_from_catalog(
        candidate: dict,
        catalog: list[dict],
        top_n: int = 3,
    ) -> list[str]:
        """基于 data/mock/channel-kb/product-catalog 的 xlsx 规则匹配。

        匹配打分:
          +2 客群标签命中 candidate tags/qualifications
          +1 行业命中 (candidate.industry 含 catalog.segment 关键词)
          +1 风控要点触达 (candidate 具备对应资质/ 专利)

        返回匹配分降序 Top-N 产品名列表(去重)。
        """
        if not catalog:
            return []
        tags = set((candidate.get("tags") or []) + (candidate.get("qualifications") or []))
        industry = candidate.get("industry", "") or ""
        scored: list[tuple[float, str]] = []
        for row in catalog:
            name = (row.get("product") or row.get("name") or "").strip()
            if not name:
                continue
            segment = (row.get("segment") or row.get("target") or "") + " " + (row.get("positioning") or "")
            risk = row.get("risk_gates") or row.get("risk") or ""
            score = 0.0
            for t in tags:
                if t and t in (segment + " " + risk):
                    score += 2.0
            if industry and any(tok in segment for tok in re.split(r"[/、，,]", industry) if tok):
                score += 1.0
            if tags and any(q in risk for q in ("专利", "专精特新", "高新")):
                score += 0.5
            if score > 0:
                scored.append((score, name))
        scored.sort(key=lambda x: x[0], reverse=True)
        out: list[str] = []
        for _, n in scored:
            if n not in out:
                out.append(n)
            if len(out) >= top_n:
                break
        return out

    @staticmethod
    def _template_pitch(hit: HitItem, ideal: IdealProfile) -> str:
        """LLM 不可用时的 fallback 话术 · 仍要含银行风格 actionable 锚点
        (per docs/contracts/agent-output-rubric-2026-05-11.md §3.1 channel)."""
        p = hit.target.payload or {}
        products = hit.extras.get("recommended_products") or []
        main_product = products[0] if products else {
            "channel_name": "专精特新企业贷",
            "category": "对公贷款",
            "max_amount": "2000万",
        }
        company = p.get("company_name", "贵司")
        industry = p.get("sub_industry") or p.get("industry") or ""
        scale = p.get("scale") or ""
        revenue = p.get("revenue_latest") or ""
        tags = p.get("tags") or []
        # 资质锚: 取 1 个最具识别度的 tag (优先含"专精特新/高新/小巨人")
        tag_priority = ("专精特新", "小巨人", "高新", "瞪羚", "上市", "国家级", "省级")
        tag_anchor = next(
            (t for t in tags for kw in tag_priority if kw in t),
            tags[0] if tags else "",
        )
        # 行业 + 规模锚 (银行 RM 视角具体化)
        biz_anchor = "、".join(filter(None, [industry, scale]))
        biz_phrase = f"贵司{biz_anchor}" if biz_anchor else company
        revenue_phrase = f"营收 {revenue} 万" if revenue else ""
        tag_phrase = f"配 {tag_anchor} 资质" if tag_anchor else ""
        context_phrase = "，".join(filter(None, [biz_phrase, revenue_phrase, tag_phrase]))
        product_name = main_product.get("channel_name", "对公贷款")
        product_amount = main_product.get("max_amount", "")
        amount_phrase = f"最高 {product_amount}" if product_amount else ""
        # 银行风格 actionable 钩子 · 含具体下一步动作 + 时限
        return (
            f"{company}您好，{context_phrase}，"
            f"匹配我们行『{product_name}』{amount_phrase}，"
            f"按同类客户经验放款周期 5-10 个工作日。"
            f"方便本周内安排 10 分钟通话沟通融资节奏吗？"
        )
