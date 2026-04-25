"""POST /api/channel/export_xlsx 冒烟测试 — 验证本地 xlsx 生成 + 字段命名合规。

关键断言：
- 端点可访问、返回 xlsx content-type
- 表头严格按 field-naming.md（enterprise_name / unified_social_credit_code / business_line /
  match_score / signal_count / signal_types / approved_amount_yuan / source_urls ...）
- 空列表输入返回 400 + VALIDATION_FAILED 错误码
"""
from __future__ import annotations

import io

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from agent_channel.api import _EXPORT_COLUMNS, app


def _candidate(name: str = "浙江测试零部件") -> dict:
    return {
        "name": name,
        "uscc": "913301020000000011",
        "registeredCapital": "1000 万",
        "founded": "2015-06-01",
        "region": "浙江",
        "industry": "精密零部件",
        "employees": 80,
        "signalScore": 55,
        "signalCount": 3,
        "signalTypes": ["bidding", "growth", "tech"],
        "signals": [
            {"type": "bidding", "title": "中标 2000 万", "detail": "",
             "date": "2026-03-01", "source": "chinabidding",
             "url": "https://chinabidding.cn/x"},
            {"type": "growth", "title": "扩产", "detail": "",
             "date": "2026-02-15", "source": "gov.cn",
             "url": "https://env.gov.cn/y"},
            {"type": "tech", "title": "专利", "detail": "",
             "date": "", "source": "cnipa", "url": ""},
        ],
        "recommendedProducts": ["流动资金贷款", "设备贷 / 固定资产贷款"],
        "dataSources": [
            {"label": "chinabidding", "hint": ""},
            {"label": "gov.cn", "hint": ""},
        ],
    }


def test_export_xlsx_returns_valid_workbook():
    client = TestClient(app)
    resp = client.post(
        "/api/channel/export_xlsx",
        json={
            "session_id": "7f4c7a2d-c2f5-4c8e-9d2a-b3e6f1a0c9d7",
            "candidates": [_candidate("浙江 A"), _candidate("浙江 B")],
            "business_line": "corporate",
        },
    )
    assert resp.status_code == 200
    assert "spreadsheetml" in resp.headers["content-type"]
    assert resp.headers["X-Agent1-Export-Rows"] == "2"

    wb = load_workbook(io.BytesIO(resp.content))
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(max_row=1))]
    # 合规：字段严格按 field-naming.md
    assert headers[:8] == [
        "enterprise_name",
        "unified_social_credit_code",
        "business_line",
        "match_score",
        "signal_count",
        "signal_types",
        "approved_amount_yuan",
        "source_urls",
    ]
    # 数据行数 = 2
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    assert len(rows) == 2
    # 第一行 enterprise_name = "浙江 A"
    assert rows[0][0] == "浙江 A"
    # business_line = corporate
    assert rows[0][2] == "corporate"
    # signal_types 是逗号拼接
    assert "bidding" in (rows[0][5] or "")
    assert "growth" in (rows[0][5] or "")
    # approved_amount_yuan 是 int 且 > 0
    assert isinstance(rows[0][6], int) and rows[0][6] > 0


def test_export_xlsx_rejects_empty():
    client = TestClient(app)
    resp = client.post(
        "/api/channel/export_xlsx",
        json={"session_id": "", "candidates": []},
    )
    assert resp.status_code == 400
    body = resp.json()
    assert body["detail"]["error"]["code"] == "VALIDATION_FAILED"


def test_export_columns_include_compliance_fields():
    """硬约束：导出列必须覆盖 onboarding §4.2 要求的核心字段。"""
    header_set = {h for h, _ in _EXPORT_COLUMNS}
    required = {
        "enterprise_name", "unified_social_credit_code", "business_line",
        "match_score", "signal_count", "signal_types",
        "approved_amount_yuan", "source_urls",
    }
    assert required.issubset(header_set)
