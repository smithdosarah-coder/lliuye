# -*- coding: utf-8 -*-
"""V16 Step 1 v2: LLM 分类结果 → review-friendly xlsx.

核心差异(v1 vs v2):
  v1: 规则预标 → 用户逐条判断对错(192 条 ≈ 15 分钟)
  v2: LLM 预标 + 规则预标双栏对比 → 用户只需重点看"分歧行"(≈ 10 分钟)

列:
  # | LLM 分类 (op+label) | 规则预标 | 一致? | LLM 置信 | 源文件 | 位置 | 文本 | LLM 理由 | 规则理由 | 你的判断 | 正确 label | 备注

排序(提高 review 效率):
  1. 分歧行 × LLM 低置信  ← 最需要人看
  2. 分歧行 × LLM 高置信
  3. 一致行 × 低置信       ← 快扫
  4. 一致行 × 高置信       ← 大概率跳过

条件格式:
  分歧行      → 整行橙色
  LLM conf<0.7→ 淡黄
  判断=✗ 错  → 浅红
  判断=⊘SKIP → 灰
"""
import json
from pathlib import Path
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter

SAMPLES_JSON = Path(__file__).parent / "outputs" / "v16_labeled_elements.json"
LLM_JSON = Path(__file__).parent / "outputs" / "v16_llm_classified.json"
OUT = Path(__file__).parent / "outputs" / "v16_REVIEW_ME_v4.xlsx"

LABEL_ORDER = ["SCAFFOLD", "PRESERVE", "FILL", "CLEAR", "SLOT", "CHECKBOX", "REWRITE"]
OP_ORDER = ["PRESERVE", "FILL", "REWRITE"]
LABEL_DESC_CH = {
    "SCAFFOLD": "骨架保留",
    "FILL":     "代码填值",
    "CLEAR":    "清空+pending",
    "REWRITE":  "LLM重写",
    "SLOT":     "占位槽",
    "CHECKBOX": "勾选框",
    "PRESERVE": "指引保留",
}

# ────────────────────────────────────────────────────────────
# Load
# ────────────────────────────────────────────────────────────
samples = json.loads(SAMPLES_JSON.read_text(encoding="utf-8"))["elements"]
llm_payload = json.loads(LLM_JSON.read_text(encoding="utf-8"))
llm_cls = llm_payload["classifications_by_location"]
meta = llm_payload["meta"]


def get_llm(loc: str) -> dict:
    return llm_cls.get(loc, {})


# ────────────────────────────────────────────────────────────
# Sort: 分歧 + 低置信 优先
# ────────────────────────────────────────────────────────────
def sort_key(s: dict) -> tuple:
    llm = get_llm(s["location"])
    llm_label = llm.get("label", "")
    rule_label = s["label"]
    disagree = 0 if (llm_label and llm_label != rule_label) else 1  # 分歧先
    llm_conf = llm.get("confidence", 0.5)
    conf_bucket = 0 if llm_conf < 0.7 else 1
    return (disagree, conf_bucket, -float(s.get("confidence", 0)))


ordered = sorted(samples, key=sort_key)

# ────────────────────────────────────────────────────────────
# Build workbook
# ────────────────────────────────────────────────────────────
wb = Workbook()
ws = wb.active
ws.title = "V16 LLM 标注 review"

headers = [
    "#", "LLM 分类", "规则预标", "一致?", "LLM置信",
    "源文件", "位置", "文本", "LLM 理由", "规则理由",
    "你的判断", "如 LLM 错,正确 label", "备注",
]
ws.append(headers)

header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill("solid", fgColor="2E5D82")
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left = Alignment(horizontal="left", vertical="center", wrap_text=True)
thin = Side(border_style="thin", color="CCCCCC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for cell in ws[1]:
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center
    cell.border = border

# 直接对"分歧行"整行硬着色(不依赖条件格式)
ORANGE = PatternFill("solid", fgColor="FFD8A8")
YELLOW = PatternFill("solid", fgColor="FFFACD")
BOTH = PatternFill("solid", fgColor="FFB466")  # 分歧 + 低置信

# 数据行
for i, s in enumerate(ordered, 1):
    row_idx = i + 1
    llm = get_llm(s["location"])
    llm_label = llm.get("label", "")
    llm_op = llm.get("op", "")
    llm_conf = llm.get("confidence", 0.0)
    llm_just = (llm.get("justification", "") or "")[:100]
    rule_label = s["label"]
    rule_just = (s.get("justification", "") or "")[:100]
    src_short = s["source"].replace("_对公成稿", "").replace("_骨架型", "")[:14]
    text = (s["text"] or "")[:180]

    disagree = (llm_label != rule_label)
    agree = "✗ 分歧" if disagree else "✓ 一致"
    low_conf = llm_conf < 0.7

    llm_cell_value = f"{llm_op} · {llm_label}" if llm_label else "(未分类)"

    ws.cell(row=row_idx, column=1, value=i)
    ws.cell(row=row_idx, column=2, value=llm_cell_value)
    ws.cell(row=row_idx, column=3, value=f"{rule_label} · {LABEL_DESC_CH.get(rule_label, '')}")
    ws.cell(row=row_idx, column=4, value=agree)
    ws.cell(row=row_idx, column=5, value=round(llm_conf, 2))
    ws.cell(row=row_idx, column=6, value=src_short)
    ws.cell(row=row_idx, column=7, value=s["location"])
    ws.cell(row=row_idx, column=8, value=text)
    ws.cell(row=row_idx, column=9, value=llm_just)
    ws.cell(row=row_idx, column=10, value=rule_just)
    ws.cell(row=row_idx, column=11, value="✓ LLM 对")  # 默认 LLM 对

    # 行底色
    if disagree and low_conf:
        row_fill = BOTH
    elif disagree:
        row_fill = ORANGE
    elif low_conf:
        row_fill = YELLOW
    else:
        row_fill = None

    for c in range(1, 14):
        cell = ws.cell(row=row_idx, column=c)
        cell.border = border
        cell.alignment = left if c in (8, 9, 10, 13) else center
        if row_fill is not None:
            cell.fill = row_fill

# 列宽
widths = {1: 5, 2: 18, 3: 16, 4: 6, 5: 8, 6: 14, 7: 14, 8: 55, 9: 35, 10: 30, 11: 14, 12: 18, 13: 18}
for idx, w in widths.items():
    ws.column_dimensions[get_column_letter(idx)].width = w

ws.freeze_panes = "A2"
for r in range(2, len(ordered) + 2):
    ws.row_dimensions[r].height = 36

# 判断下拉
dv_judge = DataValidation(
    type="list",
    formula1='"✓ LLM 对,✗ LLM 错,⊘ SKIP"',
    allow_blank=False, showDropDown=False,
)
dv_judge.prompt = "默认 LLM 对。如果 LLM 标错了改成 ✗;如果样本本身坏了改 ⊘"
dv_judge.promptTitle = "对 LLM 分类的判断"
dv_judge.add(f"K2:K{len(ordered) + 1}")
ws.add_data_validation(dv_judge)

# 正确 label 下拉
correct_labels = ",".join(LABEL_ORDER)
dv_correct = DataValidation(
    type="list",
    formula1=f'"{correct_labels}"',
    allow_blank=True, showDropDown=False,
)
dv_correct.prompt = "仅当 K 列选 ✗ LLM 错 时填,下拉选正确 label"
dv_correct.promptTitle = "正确 label"
dv_correct.add(f"L2:L{len(ordered) + 1}")
ws.add_data_validation(dv_correct)

# 分歧/低置信已经在上面硬着色;K 列的"错/SKIP"保留条件格式
light_red = PatternFill("solid", fgColor="FFDDDD")
ws.conditional_formatting.add(
    f"A2:M{len(ordered) + 1}",
    FormulaRule(formula=['$K2="✗ LLM 错"'], fill=light_red),
)
grey_fill = PatternFill("solid", fgColor="DDDDDD")
ws.conditional_formatting.add(
    f"A2:M{len(ordered) + 1}",
    FormulaRule(formula=['$K2="⊘ SKIP"'], fill=grey_fill),
)

# ────────────────────────────────────────────────────────────
# Sheet 2: 统计概览
# ────────────────────────────────────────────────────────────
disagreements = [s for s in samples if get_llm(s["location"]).get("label") != s["label"]]
low_conf = [s for s in samples if get_llm(s["location"]).get("confidence", 1.0) < 0.7]
no_cls = [s for s in samples if not get_llm(s["location"])]

ws2 = wb.create_sheet("概览")
summary = [
    ("V16 Step 1 v2 — LLM 分类 review 概览", None),
    ("", None),
    ("LLM 模型", meta.get("model", "")),
    ("Temperature", meta.get("temperature", "")),
    ("耗时", f"{meta.get('elapsed_seconds', 0)} 秒"),
    ("调用次数", meta.get("llm_stats", {}).get("call_count", 0)),
    ("总 token", meta.get("llm_stats", {}).get("total_tokens", 0)),
    ("", None),
    ("━━ 样本总览 ━━", None),
    ("总样本数", len(samples)),
    ("已分类", len(samples) - len(no_cls)),
    ("未分类(缺失)", len(no_cls)),
    ("", None),
    ("━━ 规则 vs LLM ━━", None),
    ("一致", len(samples) - len(disagreements)),
    ("分歧", len(disagreements)),
    ("一致率", f"{(1 - len(disagreements)/len(samples))*100:.1f}%"),
    ("LLM 低置信(<0.7)", len(low_conf)),
    ("", None),
    ("━━ 用户操作指引 ━━", None),
    ("1. 从第 1 行往下看(已按'分歧+低置信'优先排序)", None),
    ("2. D 列=✗ 的行是 LLM 和规则分歧的行 — 重点判断谁对", None),
    ("3. K 列默认'✓ LLM 对',如果 LLM 错了改'✗ LLM 错'并在 L 列选正确 label", None),
    ("4. 约前 113 行是分歧行,后 79 行是一致行可快扫", None),
    ("5. 判断前 60-80 行即可得到足够置信度 → 告诉我'差不多了'即可", None),
]
for i, (k, v) in enumerate(summary, 1):
    ws2.cell(row=i, column=1, value=k)
    if v is not None:
        ws2.cell(row=i, column=2, value=v)
    if i == 1:
        ws2.cell(row=i, column=1).font = Font(bold=True, size=14)
    elif str(k).startswith("━━"):
        ws2.cell(row=i, column=1).font = Font(bold=True, color="2E5D82")
ws2.column_dimensions["A"].width = 40
ws2.column_dimensions["B"].width = 30

# ────────────────────────────────────────────────────────────
# Sheet 3: Label 对齐参考
# ────────────────────────────────────────────────────────────
ws3 = wb.create_sheet("label 参考")
ref_lines = [
    ("V16 6 类 label + 3 类 op 速查", None, None),
    ("", None, None),
    ("label", "对应 op", "使用场景"),
    ("SCAFFOLD", "PRESERVE", "章节号 / 标题 / 字段标签 / 表头(例:一、基本情况 / 注册资本:)"),
    ("PRESERVE", "PRESERVE", "说明性指引(例:如涉及... / 注: / 备注:)"),
    ("FILL",     "FILL",     "有客户数据的字段取值(例:4100 万元 / 黄祖海)"),
    ("CLEAR",    "FILL",     "示例性但客户无对应(例:PD评级 / 申报单位 / 客户经理 / 绿色信贷 — 含银行方字段)"),
    ("SLOT",     "FILL",     "纯占位符(____ / 连续空格 / (面积等))"),
    ("CHECKBOX", "FILL",     "复选框字段(例:□专精特新中小企业 □小巨人 — 按客户情况勾/叉)"),
    ("REWRITE",  "REWRITE",  "正文叙述段落(例:公司主营... / 财务趋势分析...)"),
]
for i, row in enumerate(ref_lines, 1):
    for j, v in enumerate(row, 1):
        if v is not None:
            ws3.cell(row=i, column=j, value=v)
    if i == 1:
        ws3.cell(row=i, column=1).font = Font(bold=True, size=14)
    elif i == 3:
        for j in range(1, 4):
            ws3.cell(row=i, column=j).font = Font(bold=True)
            ws3.cell(row=i, column=j).fill = PatternFill("solid", fgColor="E8E8E8")
ws3.column_dimensions["A"].width = 14
ws3.column_dimensions["B"].width = 14
ws3.column_dimensions["C"].width = 60

# ────────────────────────────────────────────────────────────
# Save
# ────────────────────────────────────────────────────────────
wb.save(OUT)
print(f"[done] {OUT}")
print(f"  样本: {len(ordered)}")
print(f"  分歧: {len(disagreements)} 条(前 {len(disagreements)} 行)")
print(f"  LLM 低置信: {len(low_conf)} 条")
print(f"  一致率: {(1 - len(disagreements)/len(samples))*100:.1f}%")
